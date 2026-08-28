from email.mime import base
import os

import streamlit as st
from streamlit_option_menu import option_menu
import pandas as pd
from pathlib import Path
from utils import courbe_bam
from utils.dashboard import *
from io import BytesIO
from utils.courbe_analyse import construire_resume_courbe
import plotly.graph_objects as go
import plotly.express as px
import numpy as np
import maj



# ==========================================================
# CONFIGURATION
# ==========================================================

st.set_page_config(
    page_title="OPCVM Analytics",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ==========================================================
# CSS
# ==========================================================

st.markdown("""
<style>

.main{
    background:#f8fafc;
}

section[data-testid="stSidebar"]{
    background:#0f172a;
}

section[data-testid="stSidebar"] *{
    color:white;
}
 
div[data-testid="metric-container"]{
    background:white;
    border-radius:12px;
    padding:15px;
    box-shadow:0 2px 10px rgba(0,0,0,.08);
}

</style>
""", unsafe_allow_html=True)

# ==========================================================
# LECTURE DES DONNÉES
# ==========================================================

CHEMIN = Path("data") / "livrable1_base_opcvm.xlsx"
CHEMIN_BAM = Path("data") / "courbe_taux_BAM_2025_2026.csv"

# ==========================================================
# PIPELINE COMPLET
# ==========================================================

import utils.nettoyage as nettoyage_mod
import utils.performances as performances_mod
import utils.benchmarks as benchmarks_mod
import utils.risque as risque_mod
import utils.stress as stress_mod
import utils.scoring as scoring_mod

def safe_int(val, default="Non classé"):
    """Convertit en int en gérant les NaN."""
    return int(val) if pd.notna(val) else default
def safe_progress(score):
    """Retourne une valeur valide pour st.progress() [0.0, 1.0]."""
    if pd.notna(score):
        val = float(score) / 100
        return max(0.0, min(1.0, val))  # clamp entre 0 et 1
    return 0.0

def afficher_benchmark(fiche, label="Benchmark"):
    """
    Retourne un libellé de benchmark lisible pour une fiche (Series),
    peu importe la table d'origine (scoring, risque, reference).
    Ordre de priorité : Benchmark_Normalise -> Indice Bentchmark
    -> Indice Bentchmark_x/_y (au cas où un merge dupliqué aurait
    suffixé la colonne) -> Benchmark.
    """
    valeur = resoudre_benchmark(fiche)
    return str(valeur) if pd.notna(valeur) else "Non disponible"

def resoudre_benchmark(fiche):
    """
    Comme afficher_benchmark, mais retourne np.nan (et non la
    chaîne "Non disponible") quand rien n'est trouvé, pour les
    endroits qui font ensuite pd.notna(...) ou une comparaison
    avec table_benchmarks.index.
    """
    for colonne in [
        "Benchmark_Normalise",
        "Indice Bentchmark",
        "Indice Bentchmark_x",
        "Indice Bentchmark_y",
        "Benchmark",
    ]:
        valeur = fiche.get(colonne, np.nan)
        if pd.notna(valeur) and str(valeur).strip() not in ("", "nan", "-"):
            return valeur
    return np.nan

@st.cache_data(show_spinner="Calcul des indicateurs...")
def calculer_pipeline():

    # -----------------------
    # Chargement
    # -----------------------

    base = pd.read_excel(
        CHEMIN,
        sheet_name="Base_OPCVM"
    )
    # ← AJOUTER CECI : récupérer AN depuis les données brutes si absent
    if "AN" not in base.columns:
    # Récupère AN depuis la dernière semaine disponible dans raw/
        import glob
        dernier_fichier = max(glob.glob(str(maj.RAW / "*.xlsx")), key=os.path.getmtime)
        df_an = pd.read_excel(dernier_fichier, skiprows=1)
        df_an = df_an.rename(columns={"Dénomination OPCVM": "OPCVM"})
        an_map = df_an.set_index("CODE ISIN")["AN"].to_dict()
        base["AN"] = base["CODE ISIN"].map(an_map)

    # -----------------------
    # Nettoyage
    # -----------------------

    table_reference = nettoyage_mod.construire_table_reference(base)
    
    vl_corrigee = pd.read_excel(
        CHEMIN,
        sheet_name="VL_Historique"
    )

    vl_corrigee["Date"] = pd.to_datetime(vl_corrigee["Date"])

    vl_corrigee = vl_corrigee.set_index("Date")

    # -----------------------
    # Performances
    # -----------------------

    rendements = pd.read_excel(
        CHEMIN,
        sheet_name="Rendements_Hebdo"
    )

    rendements["Date"] = pd.to_datetime(
        rendements["Date"]
    )

    rendements = rendements.set_index("Date")

    table_performances = performances_mod.construire_table_performances(
        vl_corrigee
    )

    perf_complete = performances_mod.construire_performance_complete(
        table_reference,
        table_performances,
        vl_corrigee,
    )

    # -----------------------
    # Benchmarks
    # -----------------------
    base_benchmarks, indices_hebdo, table_perf_benchmarks = (
            benchmarks_mod.construire_benchmarks(
                table_reference,
                table_performances,
                vl_corrigee,
            )
        )
    
    # -----------------------
    # Risque
    # -----------------------
    table_risque = risque_mod.construire_risque(
        perf_complete,
        vl_corrigee,
        rendements,
    )
    # NOTE : "Indice Bentchmark" est déjà présent dans
    # table_risque via perf_complete -> table_reference,
    # donc on ne fusionne plus cette colonne ici. La faire
    # deux fois transformait la colonne en
    # "Indice Bentchmark_x" / "Indice Bentchmark_y" (suffixes
    # pandas), ce qui cassait toute lecture de la colonne
    # sous son nom original en aval.
    colonnes_a_garder = [
        "CODE ISIN",
        "Benchmark"
    ]

    colonnes_existantes = [
        c for c in colonnes_a_garder
        if c in table_reference.columns
        and c not in table_risque.columns
    ]

    if colonnes_existantes:

        table_risque = table_risque.merge(
            table_reference[["CODE ISIN"] + colonnes_existantes],
            on="CODE ISIN",
            how="left"
        )

    # -----------------------------------------------------
    # FIX : fusionner les colonnes de benchmark AVANT de
    # construire le scoring, pour que Benchmark_Normalise
    # et Benchmark_Disponible soient disponibles partout
    # en aval (table_scoring inclus). Auparavant ce merge
    # avait lieu APRES construire_scoring(), donc
    # table_scoring n'avait jamais ces colonnes et les pages
    # Recherche / Comparaison / Scoring / Classements /
    # Rapport affichaient toujours "Non disponible".
    # -----------------------------------------------------
    colonnes_benchmark = [
        "CODE ISIN",
        "Benchmark_Normalise",
        "Benchmark_Disponible",
    ]

    colonnes_benchmark = [
        c for c in colonnes_benchmark
        if c in base_benchmarks.columns
    ]

    table_risque = table_risque.merge(
        base_benchmarks[colonnes_benchmark],
        on="CODE ISIN",
        how="left",
    )

    # Le scoring est maintenant construit à partir d'une
    # table_risque qui contient déjà les colonnes benchmark.
    table_scoring = scoring_mod.construire_scoring(table_risque)

    # -----------------------------------------------------
    # FILET DE SÉCURITÉ : si construire_scoring() sélectionne
    # explicitement ses colonnes en interne (ex. df[[...]]),
    # elle peut avoir laissé tomber Benchmark_Normalise /
    # Benchmark_Disponible / Perf_1_mois / Perf_3_mois /
    # Perf_6_mois / Perf_1_an / Perf_3_ans, et même CODE ISIN,
    # même si table_risque les contient déjà. On les rattache
    # ici directement, sans dépendre du code interne de
    # utils/scoring.py. On ne tente le merge que si une clé
    # commune existe réellement des deux côtés, sinon on
    # laisse table_scoring inchangée plutôt que de planter.
    # -----------------------------------------------------
    cle_jointure = None

    for cle_candidate in ["CODE ISIN", "OPCVM"]:
        if (
            cle_candidate in table_scoring.columns
            and cle_candidate in table_risque.columns
        ):
            cle_jointure = cle_candidate
            break

    if cle_jointure is not None:

        colonnes_a_rattacher = [
            "Benchmark_Normalise",
            "Benchmark_Disponible",
            "Perf_1_mois",
            "Perf_3_mois",
            "Perf_6_mois",
            "Perf_1_an",
            "Perf_3_ans",
        ]

        colonnes_a_rattacher = [
            c for c in colonnes_a_rattacher
            if c in table_risque.columns
            and c not in table_scoring.columns
        ]

        if colonnes_a_rattacher:

            table_scoring = table_scoring.merge(
                table_risque[[cle_jointure] + colonnes_a_rattacher],
                on=cle_jointure,
                how="left",
            )

    courbe_bam = pd.read_csv(
        CHEMIN_BAM,
        parse_dates=[
            "Date_reference",
            "Date d'échéance",
            "Date de la valeur",
        ],
    )
    
    courbe_bam["Segment"] = pd.cut(
        courbe_bam["Maturite_annees"],
        bins=[0, 2, 7, 100],
        labels=[
            "Court Terme",
            "Moyen Terme",
            "Long Terme",
        ],
        include_lowest=True,
    )
    resume_courbe = construire_resume_courbe(courbe_bam)

    return {
        "table_reference": table_reference,
        "vl_corrigee": vl_corrigee,
        "rendements": rendements,
        "performances": table_performances,
        "performance_complete": perf_complete,
        "risque": table_risque,
        "scoring": table_scoring,
        "benchmarks": table_perf_benchmarks,
        "base_benchmarks": base_benchmarks,
        "indices_benchmarks": indices_hebdo,
        "courbe_bam": courbe_bam,
        "resume_courbe": resume_courbe,
    }





donnees = calculer_pipeline()

risque = donnees["risque"]


table_reference = donnees["table_reference"]

rendements = donnees["rendements"]

vl = donnees["vl_corrigee"]

performances = donnees["performances"]

performance_complete = donnees["performance_complete"]

table_risque = donnees["risque"]

table_scoring = donnees["scoring"]

courbe_bam = donnees["courbe_bam"]

resume_courbe = donnees["resume_courbe"]

table_benchmarks = donnees["benchmarks"]

base_benchmarks = donnees["base_benchmarks"]

indices_benchmarks = donnees["indices_benchmarks"]

# =====================================================
# Préparation des OPCVM obligataires
# =====================================================

opcvm_obligataires = table_reference[
    table_reference["Classification"].isin(["OCT", "OMLT"])
].copy()

opcvm_obligataires["Segment"] = (
    opcvm_obligataires["Classification"]
    .map({
        "OCT": "Court Terme",
        "OMLT": "Moyen / Long Terme",
    })
)

isin_ct = opcvm_obligataires.loc[
    opcvm_obligataires["Segment"] == "Court Terme",
    "CODE ISIN",
].tolist()

isin_mlt = opcvm_obligataires.loc[
    opcvm_obligataires["Segment"] == "Moyen / Long Terme",
    "CODE ISIN",
].tolist()

# On garde uniquement les ISIN présents dans rendements
isin_ct = [i for i in isin_ct if i in rendements.columns]
isin_mlt = [i for i in isin_mlt if i in rendements.columns]

rendement_ct = rendements[isin_ct].mean(axis=1)

rendement_mlt = rendements[isin_mlt].mean(axis=1)

performance_segments = pd.DataFrame({
    "Court Terme": rendement_ct,
    "Moyen / Long Terme": rendement_mlt,
})


# ====================================
# D3 - Evolution des performances
# ====================================

comparaison_segments = (
    opcvm_obligataires
    .groupby("Segment")
    .agg(
        Nombre_Fonds=("CODE ISIN", "count"),
        Actif_Moyen=("AN", "mean"),
        Sensibilite_Moyenne=("Sensibilité", "mean"),
    )
)

# =====================================================
# D4 - Fonds les plus exposés au risque de taux
# =====================================================

fonds_sensibles = table_risque[
    table_risque["Classification"].isin(["OCT", "OMLT"])
].copy()

fonds_sensibles = fonds_sensibles[
    [
        "OPCVM",
        "Société de Gestion",
        "Classification",
        "Sensibilité",
        "AN",
    ]
]

fonds_sensibles = fonds_sensibles.sort_values(
    "Sensibilité",
    ascending=False
)
def niveau_risque(x):

    if x >= 4:
        return "🔴 Très élevé"

    elif x >= 2:
        return "🟠 Elevé"

    elif x >= 1:
        return "🟡 Moyen"

    else:
        return "🟢 Faible"


fonds_sensibles["Niveau"] = (
    fonds_sensibles["Sensibilité"]
    .apply(niveau_risque)
)
# =====================================================
# - Impact d'un mouvement des taux
# =====================================================

impact_taux = fonds_sensibles.copy()

impact_taux["Hausse +25 pb"] = (
    -impact_taux["Sensibilité"] * 0.0025 * 100
)

impact_taux["Hausse +50 pb"] = (
    -impact_taux["Sensibilité"] * 0.0050 * 100
)

impact_taux["Hausse +100 pb"] = (
    -impact_taux["Sensibilité"] * 0.0100 * 100
)

impact_taux["Baisse -25 pb"] = (
    impact_taux["Sensibilité"] * 0.0025 * 100
)

impact_taux["Baisse -50 pb"] = (
    impact_taux["Sensibilité"] * 0.0050 * 100
)

impact_taux["Baisse -100 pb"] = (
    impact_taux["Sensibilité"] * 0.0100 * 100
)

# ==========================================================
# EXPORT EXCEL (utilisé par la page Rapport)
# ==========================================================

def generer_rapport_excel():

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        table_reference.to_excel(
            writer,
            sheet_name="Référence",
            index=False,
        )

        table_risque.to_excel(
            writer,
            sheet_name="Risque",
            index=False,
        )

        table_scoring.to_excel(
            writer,
            sheet_name="Scoring",
            index=False,
        )

        top10 = (
            table_scoring
            .sort_values("Score_Global", ascending=False)
            .head(10)
        )

        top10.to_excel(
            writer,
            sheet_name="Top10",
            index=False,
        )

    output.seek(0)

    return output

# ==========================================================
# SIDEBAR
# ==========================================================

with st.sidebar:

    st.title("📊 OPCVM Analytics")

    st.caption("Analyse quantitative des OPCVM marocains")

    st.divider()

    if st.button("🔄 Actualiser les données", key = "btn_actualiser"):

        with st.spinner("Mise à jour en cours..."):

            maj.mettre_a_jour()

        st.cache_data.clear()

        st.success("Les données ont été actualisées.")

        st.rerun()
    
    st.divider()
    
    page = option_menu(
        menu_title=None,
        options=[
            "Accueil",
            "Dashboard",
            "Recherche",
            "Comparaison",
            "Benchmarks",
            "Classements",
            "Scoring",
            "Analyse obligataire",
            "Stress Test",
            "Rapport",
            "Méthodologie",
        ],
        icons=[
            "house",
            "bar-chart",
            "search",
            "columns-gap",
            "graph-up",
            "trophy",
            "star",
            "activity",
            "lightning",
            "file-earmark-text",
            "journal-text",
        ],
        default_index=0,
    )
    st.divider()

# ==========================================================
# PAGES DE L'APPLICATION
# ==========================================================

if page == "Accueil":

    st.title("OPCVM Analytics")

    st.write(
        """
        Plateforme d'analyse quantitative des OPCVM marocains.
        """
    )

    st.divider()

    c1, c2, c3, c4 = st.columns(4)

    c1.metric(
        "Nombre de fonds",
        table_reference["CODE ISIN"].nunique()
        if not table_reference.empty else 0,
    )

    c2.metric(
        "Catégories",
        table_reference["Classification"].nunique()
        if not table_reference.empty else 0,
    )

    c3.metric(
        "Sociétés de gestion",
        table_reference["Société de Gestion"].nunique()
        if not table_reference.empty else 0,
    )

    c4.metric(
        "Historique",
        len(vl)
        if not vl.empty else 0,
    )

    st.divider()

    recherche = st.text_input(
        "🔎 Rechercher un fonds"
    )

    base = table_reference.copy()

    col1, col2 = st.columns(2)

    with col1:

        categories = ["Toutes"]

        if not base.empty:
            categories += sorted(
                base["Classification"]
                .dropna()
                .unique()
            )

        categorie = st.selectbox(
            "Classification",
            categories,
        )

    with col2:

        societes = ["Toutes"]

        if not base.empty:
            societes += sorted(
                base["Société de Gestion"]
                .dropna()
                .unique()
            )

        societe = st.selectbox(
            "Société de gestion",
            societes,
        )

    if categorie != "Toutes":
        base = base[
            base["Classification"] == categorie
        ]

    if societe != "Toutes":
        base = base[
            base["Société de Gestion"] == societe
        ]

    if recherche:

        masque = (
            base.astype(str)
            .apply(
                lambda col:
                col.str.contains(
                    recherche,
                    case=False,
                    na=False,
                )
            )
            .any(axis=1)
        )

        base = base[masque]

    st.dataframe(
        base,
        use_container_width=True,
        height=500,
    )


# ==========================================================
# DASHBOARD
# ==========================================================

elif page == "Dashboard":

    st.title("📈 Dashboard")

    # =====================================
    # Filtres
    # =====================================

    categorie = st.selectbox(
        "Filtrer par catégorie",
        ["Toutes"] + sorted(
            table_scoring["Classification"]
            .dropna()
            .unique()
        )
    )

    societe = st.selectbox(
        "Filtrer par société",
        ["Toutes"] + sorted(
            table_scoring["Société de Gestion"]
            .dropna()
            .unique()
        )
    )

    # =====================================
    # Construction des bases filtrées
    # =====================================

    base_dashboard = table_scoring.copy()
    base_risque = table_risque.copy()

    if categorie != "Toutes":

        base_dashboard = base_dashboard[
            base_dashboard["Classification"] == categorie
        ]

        base_risque = base_risque[
            base_risque["Classification"] == categorie
        ]

    if societe != "Toutes":

        base_dashboard = base_dashboard[
            base_dashboard["Société de Gestion"] == societe
        ]

        base_risque = base_risque[
            base_risque["Société de Gestion"] == societe
        ]

    # =====================================
    # KPI
    # =====================================

    afficher_kpi(base_dashboard)

    st.divider()

    col1, col2 = st.columns(2)

    with col1:
        graphique_categories(table_reference)

    with col2:
        graphique_actif(table_reference)

    st.divider()

    col1, col2 = st.columns(2)

    with col1:

        graphique_top10(base_dashboard)

        st.dataframe(
            base_dashboard
            .sort_values(
                "Score_Global",
                ascending=False
            )
            .head(10),
            use_container_width=True,
            hide_index=True,
        )

    with col2:
        graphique_performance(base_dashboard)

    st.divider()

    graphique_volatilite(base_risque)

    st.divider()

    col1, col2 = st.columns(2)

    with col1:
        graphique_pires_performances(base_risque)

    with col2:
        graphique_plus_reguliers(base_risque)

    st.dataframe(
        base_risque
        .sort_values(
            "Volatilite_annualisee_%",
            ascending=False
        )
        .head(10),
        use_container_width=True,
        hide_index=True,
    )

    st.divider()

    col1, col2 = st.columns(2)

    with col1:
        graphique_plus_volatils(base_risque)

    with col2:
        graphique_ratio(base_risque)

    st.divider()

    graphique_rendement_volatilite(base_risque)

    st.divider()

    graphique_evolution_categorie(
        rendements,
        table_reference,
    )

    st.divider()

    # =====================================
    # Répartition des fonds par benchmark
    # =====================================

    st.subheader("🔗 Répartition par benchmark")

    if "Benchmark_Normalise" in base_dashboard.columns:

        repartition_benchmark = (
            base_dashboard["Benchmark_Normalise"]
            .fillna("Non disponible")
            .value_counts()
        )

        st.bar_chart(repartition_benchmark)

    else:

        st.info(
            "Les informations de benchmark ne sont pas disponibles."
        )


# ==========================================================
# RECHERCHE
# ==========================================================

elif page == "Recherche":

    st.title("🔎 Recherche d'un OPCVM")

    liste_fonds = sorted(
        table_scoring["OPCVM"]
        .dropna()
        .unique()
    )

    fonds = st.selectbox(
        "Choisir un fonds",
        liste_fonds
    )

    fiche = table_scoring[
        table_scoring["OPCVM"] == fonds
    ].iloc[0]
    
    # =====================================
    # KPI
    # =====================================

    c1, c2, c3, c4 = st.columns(4)

    c1.metric(
        "Score global",
        f"{fiche['Score_Global']:.1f}"
    )

    c2.metric(
        "Performance YTD",
        f"{fiche['Perf_YTD_calculee_%']:.2f}%"
    )

    c3.metric(
        "Volatilité",
        f"{fiche['Volatilite_annualisee_%']:.2f}%"
    )

    c4.metric(
        "Drawdown",
        f"{fiche['Max_Drawdown_%']:.2f}%"
    )
    st.divider()

    # =====================================
    # Informations
    # =====================================

    g1, g2 = st.columns(2)

    with g1:

        st.markdown("### ℹ️ Informations")

        st.write(
            "**ISIN :**",
            fiche["CODE ISIN"]
        )

        st.write(
            "**Catégorie :**",
            fiche["Classification"]
        )

        st.write(
            "**Société :**",
            fiche["Société de Gestion"]
        )

        st.write(
            "**Benchmark :**",
            afficher_benchmark(fiche)
        )

        if "AN" in fiche.index:
            st.write(
                "**Actif Net :**",
                f"{fiche['AN']:,.0f}"
            )

    with g2:

        st.markdown("### 🏆 Classement")

        rang_cat = fiche["Rang_Categorie"]
        rang_glo = fiche["Rang_Global"]

        st.write(
            "**Rang catégorie :**",
            int(rang_cat)
            if pd.notna(rang_cat)
            else "Non classé"
        )

        st.write(
            "**Rang global :**",
            int(rang_glo)
            if pd.notna(rang_glo)
            else "Non classé"
        )

        score = fiche["Score_Global"]

        if pd.notna(score):

            st.progress(
                float(score) / 100
            )

        else:

            st.progress(0.0)

            st.caption(
                "Score non disponible"
            )

    st.divider()

    # =====================================
    # Benchmark
    # =====================================

    st.markdown("### 🔗 Benchmark")

    benchmark_normalise = afficher_benchmark(fiche)

    if benchmark_normalise != "Non disponible":

        st.write(
            "**Benchmark de référence :**",
            benchmark_normalise
        )

        if benchmark_normalise in table_benchmarks.index:

            benchmark_data = table_benchmarks.loc[
                benchmark_normalise
            ]

            comparaison = pd.DataFrame({

                "Indicateur": [
                    "Performance 1 mois",
                    "Performance 3 mois",
                    "Performance 6 mois",
                    "Performance 1 an",
                    "Performance 3 ans",
                ],

                "OPCVM": [
                    fiche.get("Perf_1_mois", np.nan),
                    fiche.get("Perf_3_mois", np.nan),
                    fiche.get("Perf_6_mois", np.nan),
                    fiche.get("Perf_1_an", np.nan),
                    fiche.get("Perf_3_ans", np.nan),
                ],

                "Benchmark": [
                    benchmark_data.get("Perf_1_mois", np.nan),
                    benchmark_data.get("Perf_3_mois", np.nan),
                    benchmark_data.get("Perf_6_mois", np.nan),
                    benchmark_data.get("Perf_1_an", np.nan),
                    benchmark_data.get("Perf_3_ans", np.nan),
                ],
            })

            comparaison["Écart"] = (
                comparaison["OPCVM"]
                - comparaison["Benchmark"]
            )

            st.dataframe(
                comparaison.style.format(
                    {
                        "OPCVM": "{:.2%}",
                        "Benchmark": "{:.2%}",
                        "Écart": "{:.2%}",
                    },
                    na_rep="N/D"
                ),
                use_container_width=True,
                hide_index=True,
            )

        else:

            st.info(
                "🔗 Les performances du benchmark ne sont pas disponibles."
            )

    else:

        st.info(
            "🔗 Aucun benchmark de référence disponible pour ce fonds."
    )
    st.divider()

    # =====================================
    # Statistiques
    # =====================================

    st.markdown("### 📋 Toutes les statistiques")

    st.dataframe(
        fiche.to_frame().T,
        use_container_width=True
    )

    st.divider()

    # =====================================
    # Evolution VL
    # =====================================

    st.subheader(
        "📉 Evolution de la Valeur Liquidative"
    )

    graphique_vl(
        vl,
        fiche["CODE ISIN"]
    )

    st.divider()


# ==========================================================
# COMPARAISON
# ==========================================================

elif page == "Comparaison":

    st.title("⇄ Comparaison de deux OPCVM")

    # =====================================
    # Sélection de la catégorie
    # =====================================

    categories = sorted(
        table_scoring["Classification"]
        .dropna()
        .astype(str)
        .unique()
    )

    categorie_choisie = st.selectbox(
        "Catégorie",
        ["Toutes les catégories"] + categories,
        key="categorie_comparaison"
    )

    # =====================================
    # Filtrage des fonds selon la catégorie
    # =====================================

    if categorie_choisie == "Toutes les catégories":

        fonds = sorted(
            table_scoring["OPCVM"]
            .dropna()
            .astype(str)
            .unique()
        )

    else:

        fonds = sorted(
            table_scoring.loc[
                table_scoring["Classification"].astype(str)
                == categorie_choisie,
                "OPCVM"
            ]
            .dropna()
            .astype(str)
            .unique()
        )

    # Vérification du nombre de fonds disponibles

    if len(fonds) < 2:

        st.warning(
            "⚠️ Cette catégorie ne contient pas suffisamment de fonds "
            "pour effectuer une comparaison."
        )

        st.stop()

    # =====================================
    # Sélection des deux fonds
    # =====================================

    col1, col2 = st.columns(2)

    with col1:

        fonds1 = st.selectbox(
            "Premier fonds",
            fonds,
            index=0,
            key="fonds1",
        )

    with col2:

        fonds2 = st.selectbox(
            "Deuxième fonds",
            fonds,
            index=1,
            key="fonds2",
        )

    # =====================================
    # Récupération des fiches
    # =====================================

    fiche1 = table_scoring[
        table_scoring["OPCVM"] == fonds1
    ].iloc[0]

    fiche2 = table_scoring[
        table_scoring["OPCVM"] == fonds2
    ].iloc[0]
    st.divider()

    # =====================================
    # Comparaison principale
    # =====================================

    comparaison = pd.DataFrame({

        "Indicateur": [
            "Classification",
            "Benchmark",
            "Actif Net",
            "Performance YTD (%)",
            "Volatilité (%)",
            "Max Drawdown (%)",
            "Score Global",
            "Rang Global",
        ],

        fonds1: [

            fiche1["Classification"],

            afficher_benchmark(fiche1),

            f"{fiche1['AN']:,.0f}",

            round(
                fiche1["Perf_YTD_calculee_%"],
                2
            ),

            round(
                fiche1["Volatilite_annualisee_%"],
                2
            ),

            round(
                fiche1["Max_Drawdown_%"],
                2
            ),

            round(
                fiche1["Score_Global"],
                2
            ),

            "-"
            if pd.isna(fiche1["Rang_Global"])
            else int(fiche1["Rang_Global"])
        ],

        fonds2: [

            fiche2["Classification"],

            afficher_benchmark(fiche2),

            f"{fiche2['AN']:,.0f}",

            round(
                fiche2["Perf_YTD_calculee_%"],
                2
            ),

            round(
                fiche2["Volatilite_annualisee_%"],
                2
            ),

            round(
                fiche2["Max_Drawdown_%"],
                2
            ),

            round(
                fiche2["Score_Global"],
                2
            ),

            "-"
            if pd.isna(fiche2["Rang_Global"])
            else int(fiche2["Rang_Global"]),
        ],
    })

    st.subheader(
        "📋 Comparaison des indicateurs"
    )

    st.dataframe(
        comparaison,
        use_container_width=True,
        hide_index=True,
    )

    # =====================================
    # Comparaison avec les benchmarks
    # =====================================

    st.divider()

    st.subheader(
        "🔗 Comparaison avec les benchmarks"
    )

    benchmark1 = resoudre_benchmark(fiche1)
    benchmark2 = resoudre_benchmark(fiche2)

    col1, col2 = st.columns(2)

    with col1:

        st.markdown(
            f"**{fonds1}**"
        )

        st.write(
            "Benchmark :",
            benchmark1
            if pd.notna(benchmark1)
            else "Non disponible"
        )

    with col2:

        st.markdown(
            f"**{fonds2}**"
        )

        st.write(
            "Benchmark :",
            benchmark2
            if pd.notna(benchmark2)
            else "Non disponible"
        )

    # Comparaison seulement si les deux fonds
    # possèdent un benchmark disponible

    if (
        pd.notna(benchmark1)
        and pd.notna(benchmark2)
        and benchmark1 in table_benchmarks.index
        and benchmark2 in table_benchmarks.index
    ):

        periodes = [
            "Perf_1_mois",
            "Perf_3_mois",
            "Perf_6_mois",
            "Perf_1_an",
        ]

        labels = [
            "Performance 1 mois",
            "Performance 3 mois",
            "Performance 6 mois",
            "Performance 1 an",
        ]

        comparaison_benchmarks = pd.DataFrame({

            "Indicateur": labels,

            fonds1: [
                fiche1.get(
                    "Perf_1_mois",
                    np.nan
                ),
                fiche1.get(
                    "Perf_3_mois",
                    np.nan
                ),
                fiche1.get(
                    "Perf_6_mois",
                    np.nan
                ),
                fiche1.get(
                    "Perf_1_an",
                    np.nan
                ),
            ],

            f"Benchmark {benchmark1}": [
                table_benchmarks.loc[
                    benchmark1,
                    p
                ]
                for p in periodes
            ],

            fonds2: [
                fiche2.get(
                    "Perf_1_mois",
                    np.nan
                ),
                fiche2.get(
                    "Perf_3_mois",
                    np.nan
                ),
                fiche2.get(
                    "Perf_6_mois",
                    np.nan
                ),
                fiche2.get(
                    "Perf_1_an",
                    np.nan
                ),
            ],

            f"Benchmark {benchmark2}": [
                table_benchmarks.loc[
                    benchmark2,
                    p
                ]
                for p in periodes
            ],
        })

        st.dataframe(
            comparaison_benchmarks.style.format(
                {
                    fonds1: "{:.2%}",
                    f"Benchmark {benchmark1}": "{:.2%}",
                    fonds2: "{:.2%}",
                    f"Benchmark {benchmark2}": "{:.2%}",
                }
            ),
            use_container_width=True,
            hide_index=True,
        )

    else:

        st.info(
            "🔗 Les benchmarks nécessaires à la comparaison ne sont pas disponibles."
        )


# ==========================================================
# CLASSEMENTS
# ==========================================================

elif page == "Classements":

    st.title("🏆 Classements des OPCVM")

    col1, col2, col3 = st.columns(3)

    with col1:

        categorie = st.selectbox(
            "Catégorie",
            ["Toutes"]
            + sorted(
                table_scoring[
                    "Classification"
                ]
                .dropna()
                .unique()
            )
        )

    with col2:

        tri = st.selectbox(
            "Classer selon",
            [
                "Score_Global",
                "Perf_YTD_calculee_%",
                "Volatilite_annualisee_%",
                "Max_Drawdown_%"
            ]
        )

    with col3:

        top = st.slider(
            "Nombre de fonds",
            5,
            50,
            20
        )

    classement = table_scoring.copy()

    if categorie != "Toutes":

        classement = classement[
            classement["Classification"]
            == categorie
        ]

    if tri in [
        "Volatilite_annualisee_%",
        "Max_Drawdown_%"
    ]:

        asc = True

    else:

        asc = False

    classement = classement.sort_values(
        tri,
        ascending=asc
    ).head(top)

    colonnes_classement = [
        "Rang_Global",
        "OPCVM",
        "Classification",
        "Score_Global",
        "Perf_YTD_calculee_%",
        "Volatilite_annualisee_%",
        "Max_Drawdown_%"
    ]

    # Ajouter le benchmark s'il existe
    if "Benchmark_Normalise" in classement.columns:
        colonnes_classement.insert(
            3,
            "Benchmark_Normalise"
        )

    st.dataframe(
        classement[
            colonnes_classement
        ],
        use_container_width=True,
        hide_index=True,
    )

    st.download_button(
        "⬇️ Télécharger le classement",
        classement.to_csv(
            index=False
        ).encode("utf-8-sig"),
        "classement_opcvm.csv",
        "text/csv"
    )


# ==========================================================
# SCORING
# ==========================================================

elif page == "Scoring":

    st.title("⭐ Scoring des OPCVM")

    st.markdown(
        """
        Le score global est calculé sur **100 points** à partir de :

        - Performance
        - Volatilité
        - Régularité
        - Drawdown
        - Actif Net
        """
    )

    
    st.divider()

    col1, col2 = st.columns([1, 2])

    with col1:

        fonds = st.selectbox(
            "Choisir un fonds",
            sorted(
                table_scoring[
                    "OPCVM"
                ]
                .dropna()
                .unique()
            )
        )

        fiche = table_scoring[
            table_scoring["OPCVM"] == fonds
        ].iloc[0]

        st.metric(
            "Score Global",
            f"{fiche['Score_Global']:.1f}"
        )

        st.metric(
            "Rang",
            int(fiche["Rang_Global"])
        )

        st.caption(
            f"🔗 Benchmark : {afficher_benchmark(fiche)}"
        )

        st.progress(
            safe_progress(
                fiche["Score_Global"]
            )
        )

        st.write(
            "### Détail du score"
        )

        st.progress(
            safe_progress(
                fiche["Score_Performance"]
            )
        )

        st.caption(
            f"Performance : "
            f"{fiche['Score_Performance']:.1f}"
        )

        st.progress(
            safe_progress(
                fiche["Score_Volatilite"]
            )
        )

        st.caption(
            f"Volatilité : "
            f"{fiche['Score_Volatilite']:.1f}"
        )

        st.progress(
            safe_progress(
                fiche["Score_Regularite"]
            )
        )

        st.caption(
            f"Régularité : "
            f"{fiche['Score_Regularite']:.1f}"
        )

        st.progress(
            safe_progress(
                fiche["Score_Drawdown"]
            )
        )

        st.caption(
            f"Drawdown : "
            f"{fiche['Score_Drawdown']:.1f}"
        )

        st.progress(
            safe_progress(
                fiche["Score_Taille"]
            )
        )

        st.caption(
            f"Taille : "
            f"{fiche['Score_Taille']:.1f}"
        )

    with col2:

        scores = pd.DataFrame({

            "Critère": [
                "Performance",
                "Volatilité",
                "Régularité",
                "Drawdown",
                "Taille"
            ],

            "Score": [
                fiche["Score_Performance"],
                fiche["Score_Volatilite"],
                fiche["Score_Regularite"],
                fiche["Score_Drawdown"],
                fiche["Score_Taille"],
            ]

        })

        st.subheader(
            "📊 Répartition du score"
        )

        st.bar_chart(
            scores.set_index("Critère"),
            use_container_width=True,
        )

    st.divider()

    st.subheader(
        "📊 Distribution des scores"
    )

    score = (
        table_scoring["Score_Global"]
        .round(0)
        .value_counts()
        .sort_index()
    )

    st.bar_chart(score)

    st.subheader(
        "🏆 Top 10 des OPCVM"
    )

    top10 = (
        table_scoring
        .sort_values(
            "Score_Global",
            ascending=False
        )
        .head(10)
    )

    colonnes_top10 = [
        "Rang_Global",
        "OPCVM",
        "Classification",
        "Score_Global",
    ]

    if "Benchmark_Normalise" in top10.columns:
        colonnes_top10.insert(
            3,
            "Benchmark_Normalise"
        )

    st.dataframe(
        top10[
            colonnes_top10
        ],
        use_container_width=True,
        hide_index=True,
    )

    st.divider()

    st.subheader(
        "📊 Répartition des catégories dans le Top 20"
    )

    top20 = (
        table_scoring
        .sort_values(
            "Score_Global",
            ascending=False
        )
        .head(20)
    )

    st.bar_chart(
        top20["Classification"].value_counts()
    )

    st.subheader(
        "📋 Tous les scores"
    )

    st.dataframe(
        table_scoring.sort_values(
            "Score_Global",
            ascending=False,
        ),
        use_container_width=True,
        hide_index=True,
    )


# ==========================================================
# BENCHMARKS
# ==========================================================

elif page == "Benchmarks":

    st.title("🔗 Benchmarks")

    st.markdown(
        "Comparaison des OPCVM avec leur indice ou taux de référence."
    )

    # =====================================
    # Performances des benchmarks
    # =====================================

    st.subheader(
        "📊 Performances des benchmarks"
    )

    bench_affichage = (
        table_benchmarks
        .reset_index()
        .rename(
            columns={
                "Indice": "Benchmark"
            }
        )
    )

    st.dataframe(
        bench_affichage.style.format(
            {
                col: "{:.2%}"
                for col in bench_affichage.columns
                if col != "Benchmark"
            }
        ),
        use_container_width=True,
        hide_index=True,
    )

    st.divider()

    # =====================================
    # Choix d'un fonds
    # =====================================

    fonds = sorted(
        table_scoring["OPCVM"]
        .dropna()
        .unique()
    )

    fonds_selectionne = st.selectbox(
        "Choisir un OPCVM",
        fonds,
    )

    fiche = table_risque[
        table_risque["OPCVM"]
        == fonds_selectionne
    ].iloc[0]

    benchmark = resoudre_benchmark(fiche)

    st.write(
        "**Benchmark de référence :**",
        benchmark
        if pd.notna(benchmark)
        else "Non disponible"
    )

    # =====================================
    # Comparaison
    # =====================================

    if (
        pd.notna(benchmark)
        and benchmark in table_benchmarks.index
    ):

        comparaison_benchmark = pd.DataFrame({

            "Indicateur": [
                "Performance 1 mois",
                "Performance 3 mois",
                "Performance 6 mois",
                "Performance 1 an",
            ],

            "OPCVM": [
                fiche.get(
                    "Perf_1_mois",
                    np.nan
                ),
                fiche.get(
                    "Perf_3_mois",
                    np.nan
                ),
                fiche.get(
                    "Perf_6_mois",
                    np.nan
                ),
                fiche.get(
                    "Perf_1_an",
                    np.nan
                ),
            ],

            "Benchmark": [
                table_benchmarks.loc[
                    benchmark,
                    "Perf_1_mois"
                ],
                table_benchmarks.loc[
                    benchmark,
                    "Perf_3_mois"
                ],
                table_benchmarks.loc[
                    benchmark,
                    "Perf_6_mois"
                ],
                table_benchmarks.loc[
                    benchmark,
                    "Perf_1_an"
                ],
            ],
        })

        comparaison_benchmark["Écart"] = (
            comparaison_benchmark["OPCVM"]
            - comparaison_benchmark["Benchmark"]
        )

        st.subheader(
            "📋 Comparaison OPCVM / Benchmark"
        )

        st.dataframe(
            comparaison_benchmark.style.format(
                {
                    "OPCVM": "{:.2%}",
                    "Benchmark": "{:.2%}",
                    "Écart": "{:.2%}",
                }
            ),
            use_container_width=True,
            hide_index=True,
        )

    else:

        st.info(
            "🔗 Aucun benchmark disponible pour ce fonds."
        )


# ==========================================================
# ANALYSE OBLIGATAIRE
# ==========================================================

elif page == "Analyse obligataire":

    st.title("📉 Analyse des OPCVM obligataires")

    st.markdown(
        """
        Cette page analyse la courbe des taux des Bons du Trésor publiée par
        **Bank Al-Maghrib** et son impact potentiel sur les OPCVM obligataires.
        """
    )

    st.divider()

    # ==========================================================
    # COURBE DES TAUX
    # ==========================================================

    dates = sorted(
        courbe_bam["Date_reference"].unique()
    )

    date_selectionnee = st.selectbox(
        "Date de la courbe",
        dates,
        index=len(dates) - 1,
    )

    courbe = courbe_bam[
        courbe_bam["Date_reference"] == date_selectionnee
    ].copy()

    # Conversion numérique
    courbe["Maturite_annees"] = pd.to_numeric(
        courbe["Maturite_annees"],
        errors="coerce"
    )

    courbe["Taux"] = pd.to_numeric(
        courbe["Taux"],
        errors="coerce"
    )

    courbe = courbe.dropna(
        subset=[
            "Maturite_annees",
            "Taux"
        ]
    )

    courbe = courbe.sort_values(
        "Maturite_annees"
    )

    st.subheader(
        "📈 Courbe des taux"
    )

    st.line_chart(
        data=courbe,
        x="Maturite_annees",
        y="Taux",
    )

    # ==========================================================
    # TAUX MOYENS PAR SEGMENT
    # ==========================================================

    st.subheader(
        "📊 Taux moyens"
    )

    ct = courbe[
        courbe["Segment"] == "Court Terme"
    ]["Taux"].mean()

    mt = courbe[
        courbe["Segment"] == "Moyen Terme"
    ]["Taux"].mean()

    lt = courbe[
        courbe["Segment"] == "Long Terme"
    ]["Taux"].mean()

    c1, c2, c3 = st.columns(3)

    c1.metric(
        "Court Terme",
        f"{ct:.2f}%"
    )

    c2.metric(
        "Moyen Terme",
        f"{mt:.2f}%"
    )

    c3.metric(
        "Long Terme",
        f"{lt:.2f}%"
    )

    st.divider()

    # ==========================================================
    # COMPARAISON DE DEUX COURBES
    # ==========================================================

    st.subheader(
        "📈 Comparaison de deux courbes des taux"
    )

    col1, col2 = st.columns(2)

    dates = sorted(
        courbe_bam[
            "Date_reference"
        ].drop_duplicates()
    )

    with col1:

        date1 = st.selectbox(
            "Courbe 1",
            dates,
            index=max(
                0,
                len(dates) - 10
            ),
            key="date_courbe_1"
        )

    with col2:

        date2 = st.selectbox(
            "Courbe 2",
            dates,
            index=len(dates) - 1,
            key="date_courbe_2"
        )

    # ==========================================================
    # EXTRACTION DES DEUX COURBES
    # ==========================================================

    courbe1 = courbe_bam[
        courbe_bam["Date_reference"] == date1
    ].copy()

    courbe2 = courbe_bam[
        courbe_bam["Date_reference"] == date2
    ].copy()

    # ==========================================================
    # CONVERSION NUMÉRIQUE
    # ==========================================================

    courbe1["Maturite_annees"] = pd.to_numeric(
        courbe1["Maturite_annees"],
        errors="coerce"
    )

    courbe1["Taux"] = pd.to_numeric(
        courbe1["Taux"],
        errors="coerce"
    )

    courbe2["Maturite_annees"] = pd.to_numeric(
        courbe2["Maturite_annees"],
        errors="coerce"
    )

    courbe2["Taux"] = pd.to_numeric(
        courbe2["Taux"],
        errors="coerce"
    )

    # Suppression des observations invalides
    courbe1 = courbe1.dropna(
        subset=[
            "Maturite_annees",
            "Taux"
        ]
    )

    courbe2 = courbe2.dropna(
        subset=[
            "Maturite_annees",
            "Taux"
        ]
    )

    # Tri
    courbe1 = courbe1.sort_values(
        "Maturite_annees"
    )

    courbe2 = courbe2.sort_values(
        "Maturite_annees"
    )

    # ==========================================================
    # GRAPHIQUE COMPARATIF
    # ==========================================================

    fig = go.Figure()

    fig.add_trace(
        go.Scatter(
            x=courbe1["Maturite_annees"],
            y=courbe1["Taux"],
            mode="lines+markers",
            name=date1.strftime("%d/%m/%Y"),
            line=dict(width=3)
        )
    )

    fig.add_trace(
        go.Scatter(
            x=courbe2["Maturite_annees"],
            y=courbe2["Taux"],
            mode="lines+markers",
            name=date2.strftime("%d/%m/%Y"),
            line=dict(width=3)
        )
    )

    fig.update_layout(
        template="plotly_white",
        height=550,
        hovermode="x unified",
        legend=dict(
            orientation="h",
            y=1.08,
            x=0
        ),
        title="Comparaison de la structure des taux",
        xaxis_title="Maturité (années)",
        yaxis_title="Taux (%)",
        margin=dict(
            l=40,
            r=30,
            t=70,
            b=40
        )
    )

    st.plotly_chart(
        fig,
        use_container_width=True,
    )

    # ==========================================================
    # COMPARAISON NUMÉRIQUE DES COURBES
    # ==========================================================

    # Grille commune de maturités.
    # Les courbes BAM ne possèdent pas nécessairement
    # exactement les mêmes maturités selon les dates.
    maturites_standard = np.array([
        0.25,
        0.50,
        1,
        2,
        3,
        5,
        7,
        10,
        15,
        20,
        30
    ])

    # Bornes disponibles pour chaque courbe
    x1 = courbe1["Maturite_annees"].to_numpy()
    y1 = courbe1["Taux"].to_numpy()

    x2 = courbe2["Maturite_annees"].to_numpy()
    y2 = courbe2["Taux"].to_numpy()

    if len(x1) < 2 or len(x2) < 2:

        st.warning(
            "⚠️ Données insuffisantes pour comparer les deux courbes."
        )

    else:

        # ======================================================
        # SÉLECTION DES MATURITÉS COMMUNES
        # ======================================================

        grille1 = maturites_standard[
            (maturites_standard >= x1.min())
            &
            (maturites_standard <= x1.max())
        ]

        grille2 = maturites_standard[
            (maturites_standard >= x2.min())
            &
            (maturites_standard <= x2.max())
        ]

        grille = np.intersect1d(
            grille1,
            grille2
        )

        if len(grille) < 2:

            st.warning(
                "⚠️ Les deux courbes ne couvrent pas suffisamment "
                "de maturités communes pour effectuer la comparaison."
            )

        else:

            # ==================================================
            # INTERPOLATION
            # ==================================================

            taux1_interp = np.interp(
                grille,
                x1,
                y1
            )

            taux2_interp = np.interp(
                grille,
                x2,
                y2
            )

            # ==================================================
            # TABLEAU DE COMPARAISON
            # ==================================================

            comparaison = pd.DataFrame({

                "Maturite_annees": grille,

                "Taux_1": taux1_interp,

                "Taux_2": taux2_interp

            })

            # ==================================================
            # VARIATION
            # ==================================================

            comparaison["Variation"] = (
                comparaison["Taux_2"]
                -
                comparaison["Taux_1"]
            )

            # ==================================================
            # ANALYSE DU DÉPLACEMENT
            # ==================================================

            variation = comparaison[
                "Variation"
            ].mean()

            # Court terme : maturités <= 2 ans
            court = comparaison.loc[
                comparaison["Maturite_annees"] <= 2,
                "Variation"
            ].mean()

            # Long terme : maturités >= 7 ans
            long = comparaison.loc[
                comparaison["Maturite_annees"] >= 7,
                "Variation"
            ].mean()

            # Différence de déplacement
            pente = long - court

            # ==================================================
            # AFFICHAGE DES INDICATEURS
            # ==================================================

            st.subheader(
                "📐 Analyse du déplacement"
            )

            col1, col2, col3 = st.columns(3)

            col1.metric(
                "Variation moyenne",
                f"{variation:.2f} pts"
            )

            col2.metric(
                "Variation LT - CT",
                f"{pente:.2f} pts"
            )

            # ==================================================
            # CLASSIFICATION DU MOUVEMENT
            # ==================================================

            if pente > 0.15:

                interpretation = (
                    "Steepening (pentification)"
                )

            elif pente < -0.15:

                interpretation = (
                    "Flattening (aplatissement)"
                )

            else:

                interpretation = (
                    "Déplacement parallèle"
                )

            col3.metric(
                "Type de mouvement",
                interpretation
            )

            # ==================================================
            # COMMENTAIRE SUR LE DÉPLACEMENT
            # ==================================================

            st.markdown(
                "### 💬 Commentaire"
            )

            variation_pb = variation * 100
            court_pb = court * 100
            long_pb = long * 100
            pente_pb = pente * 100


            if interpretation == "Steepening (pentification)":

                st.warning(
                    f"""
                    **Pentification de la courbe :** les taux longs ont augmenté
                        davantage que les taux courts.

                    • Variation CT : **{court_pb:.1f} pb**
                    • Variation LT : **{long_pb:.1f} pb**
                    • Écart LT - CT : **{pente_pb:.1f} pb**
                    • Variation moyenne : **{variation_pb:.1f} pb**
                    """
                )


            elif interpretation == "Flattening (aplatissement)":

                st.info(
                    f"""
                    **Aplatissement de la courbe :** les taux courts ont évolué
                    davantage que les taux longs.

                    • Variation CT : **{court_pb:.1f} pb**
                    • Variation LT : **{long_pb:.1f} pb**
                    • Écart LT - CT : **{pente_pb:.1f} pb**
                    • Variation moyenne : **{variation_pb:.1f} pb**
                    """
                )



            else:

                st.success(
                    f"""
                    **Déplacement quasiment parallèle :** les taux ont évolué
                    de manière relativement homogène sur les différentes maturités.

                    • Variation CT : **{court_pb:.1f} pb**
                    • Variation LT : **{long_pb:.1f} pb**
                    • Écart LT - CT : **{pente_pb:.1f} pb**
                    • Variation moyenne : **{variation_pb:.1f} pb**
                    """
                )
            # ==================================================
            # TAUX CT / LT DE LA COURBE 2
            # ==================================================

            col1, col2 = st.columns(2)

            col1.metric(
                "Taux CT",
                f"{courbe2.iloc[0]['Taux']:.2f}%"
            )

            col2.metric(
                "Taux LT",
                f"{courbe2.iloc[-1]['Taux']:.2f}%"
            )

            # ==================================================
            # INTERPRÉTATION DE LA VARIATION MOYENNE
            # ==================================================

            if variation > 0.20:

                st.error(
                    "⚠️ Déplacement haussier significatif de la courbe des taux."
                )

            elif variation > 0:

                st.warning(
                    "⚠️ Hausse modérée des taux."
                )

            elif variation < -0.20:

                st.success(
                    "✅ Baisse significative des taux."
                )

            else:

                st.info(
                    "ℹ️ Courbe globalement stable."
                )

    # ==========================================================
    # COMPARAISON DES SEGMENTS OBLIGATAIRES
    # ==========================================================

    st.subheader(
        "📊 Comparaison des segments obligataires"
    )

    st.dataframe(
        comparaison_segments.style.format({
            "Actif_Moyen": "{:,.0f}",
            "Sensibilite_Moyenne": "{:.2f}",
            "Performance_Moyenne": "{:.2f} %",
        }),
        use_container_width=True,
    )

    fig = px.bar(
        comparaison_segments.reset_index(),
        x="Segment",
        y="Sensibilite_Moyenne",
        color="Segment",
        text_auto=".2f",
        title="Sensibilité moyenne par segment",
    )

    fig.update_layout(
        xaxis_title="",
        yaxis_title="Sensibilité",
        height=450,
    )

    st.plotly_chart(
        fig,
        use_container_width=True,
    )

    # ==========================================================
    #• FONDS LES PLUS EXPOSÉS AU RISQUE DE TAUX
    # ==========================================================

    st.header(
        "📊 • Fonds les plus exposés au risque de taux"
    )

    st.dataframe(
        fonds_sensibles,
        use_container_width=True,
        hide_index=True,
    )

    top20 = fonds_sensibles.head(20)

    fig = px.bar(
        top20,
        x="Sensibilité",
        y="OPCVM",
        orientation="h",
        color="Sensibilité",
        title="Top 20 des OPCVM les plus sensibles aux taux",
    )

    fig.update_layout(
        yaxis={
            "categoryorder": "total ascending"
        },
        height=700,
    )

    st.plotly_chart(
        fig,
        use_container_width=True,
    )

    # ==========================================================
    #  • IMPACT POTENTIEL DES MOUVEMENTS DE TAUX
    # ==========================================================

    st.header(
        "📊 • Impact potentiel des mouvements de taux"
    )

    st.dataframe(
        impact_taux.style.format({
            "Hausse +25 pb": "{:.2f} %",
            "Hausse +50 pb": "{:.2f} %",
            "Hausse +100 pb": "{:.2f} %",
            "Baisse -25 pb": "{:.2f} %",
            "Baisse -50 pb": "{:.2f} %",
            "Baisse -100 pb": "{:.2f} %",
        }),
        use_container_width=True
    )

    top10 = impact_taux.head(10)

    fig = px.bar(
        top10,
        x="OPCVM",
        y="Hausse +100 pb",
        color="Hausse +100 pb",
        title="Impact estimé d'une hausse de 100 pb",
    )

    st.plotly_chart(
        fig,
        use_container_width=True
    )

    # ==========================================================
    # D6 • STRESS TESTING
    # ==========================================================

    st.header(
        "⚡ • Stress Testing"
    )

    scenario = st.selectbox(
        "Scénario",
        [
            "+25 pb",
            "+50 pb",
            "+100 pb",
            "Aplatissement",
            "Pentification",
        ]
    )

    stress = fonds_sensibles.copy()

    if scenario == "+25 pb":

        stress["Impact"] = (
            -stress["Sensibilité"]
            * 0.0025
            * 100
        )

    elif scenario == "+50 pb":

        stress["Impact"] = (
            -stress["Sensibilité"]
            * 0.0050
            * 100
        )

    elif scenario == "+100 pb":

        stress["Impact"] = (
            -stress["Sensibilité"]
            * 0.0100
            * 100
        )

    elif scenario == "Aplatissement":

        stress["Impact"] = np.where(
            stress["Classification"] == "OCT",
            -stress["Sensibilité"]
            * 0.0050
            * 100,
            0,
        )

    elif scenario == "Pentification":

        stress["Impact"] = np.where(
            stress["Classification"] == "OMLT",
            -stress["Sensibilité"]
            * 0.0050
            * 100,
            0,
        )

    stress = stress.sort_values(
        "Impact"
    )

    st.dataframe(
        stress,
        use_container_width=True,
        hide_index=True,
    )

    fig = px.bar(
        stress.head(20),
        x="Impact",
        y="OPCVM",
        orientation="h",
        color="Impact",
        title=f"Stress Test : {scenario}",
    )

    st.plotly_chart(
        fig,
        use_container_width=True
    )

# ==========================================================
# STRESS TEST
# ==========================================================

elif page == "Stress Test":

    st.title("⚡ Stress Testing des OPCVM obligataires")

    st.markdown(
        """
        Cette simulation estime l'impact potentiel des mouvements
        de la courbe des taux sur la valeur d'un OPCVM obligataire,
        à partir de sa sensibilité au risque de taux.
        """
    )

    st.divider()

    # ==========================================================
    # OPCVM ÉLIGIBLES
    # ==========================================================

    fonds_obligataires = table_risque.copy()

    fonds_obligataires["Sensibilité"] = pd.to_numeric(
        fonds_obligataires["Sensibilité"],
        errors="coerce"
    )

    fonds_obligataires = fonds_obligataires.dropna(
        subset=["Sensibilité"]
    )

    fonds_obligataires = fonds_obligataires[
        fonds_obligataires["Classification"].isin(
            [
                "OMLT",
                "OCT",
                "Monétaire"
            ]
        )
    ]

    # ==========================================================
    # CHOIX DU FONDS
    # ==========================================================

    fonds = sorted(
        fonds_obligataires["OPCVM"]
        .dropna()
        .astype(str)
        .unique()
    )

    if len(fonds) == 0:

        st.warning(
            "Aucun OPCVM avec une sensibilité disponible."
        )

        st.stop()

    fonds_selectionne = st.selectbox(
        "Choisir un OPCVM obligataire",
        fonds
    )

    fiche = fonds_obligataires[
        fonds_obligataires["OPCVM"]
        == fonds_selectionne
    ].iloc[0]

    sensibilite = fiche["Sensibilité"]

    # ==========================================================
    # INFORMATIONS DU FONDS
    # ==========================================================

    st.write(
        fiche[
            [
                "OPCVM",
                "Classification",
                "Sensibilité",
            ]
        ]
    )

    st.divider()

    # ==========================================================
    # CHOIX DU SCÉNARIO
    # ==========================================================

    scenario = st.selectbox(
        "Choisir un scénario de stress",
        [
            "Hausse +25 pb",
            "Hausse +50 pb",
            "Hausse +100 pb",
            "Baisse -25 pb",
            "Baisse -50 pb",
            "Baisse -100 pb",
            "Aplatissement",
            "Pentification",
        ]
    )

    # ==========================================================
    # CALCUL DU SCÉNARIO
    # ==========================================================

    if scenario == "Hausse +25 pb":

        variation_taux = 25 / 10000

        impact = (
            -sensibilite
            * variation_taux
            * 100
        )

        scenario_pb = "+25 pb"

        interpretation_scenario = (
            "Hausse uniforme de 25 points de base des taux."
        )

    elif scenario == "Hausse +50 pb":

        variation_taux = 50 / 10000

        impact = (
            -sensibilite
            * variation_taux
            * 100
        )

        scenario_pb = "+50 pb"

        interpretation_scenario = (
            "Hausse uniforme de 50 points de base des taux."
        )

    elif scenario == "Hausse +100 pb":

        variation_taux = 100 / 10000

        impact = (
            -sensibilite
            * variation_taux
            * 100
        )

        scenario_pb = "+100 pb"

        interpretation_scenario = (
            "Hausse uniforme de 100 points de base des taux."
        )

    elif scenario == "Baisse -25 pb":

        variation_taux = -25 / 10000

        impact = (
            -sensibilite
            * variation_taux
            * 100
        )

        scenario_pb = "-25 pb"

        interpretation_scenario = (
            "Baisse uniforme de 25 points de base des taux."
        )

    elif scenario == "Baisse -50 pb":

        variation_taux = -50 / 10000

        impact = (
            -sensibilite
            * variation_taux
            * 100
        )

        scenario_pb = "-50 pb"

        interpretation_scenario = (
            "Baisse uniforme de 50 points de base des taux."
        )

    elif scenario == "Baisse -100 pb":

        variation_taux = -100 / 10000

        impact = (
            -sensibilite
            * variation_taux
            * 100
        )

        scenario_pb = "-100 pb"

        interpretation_scenario = (
            "Baisse uniforme de 100 points de base des taux."
        )

    # ==========================================================
    # APLATISSEMENT
    # ==========================================================

    elif scenario == "Aplatissement":

        # Hypothèse simplifiée :
        # hausse de 50 pb sur la partie courte
        # et stabilité de la partie longue.

        variation_taux = 50 / 10000

        impact = (
            -sensibilite
            * variation_taux
            * 100
        )

        scenario_pb = "Scénario non parallèle"

        interpretation_scenario = (
            "Hausse de la partie courte de la courbe "
            "avec stabilité relative de la partie longue."
        )

    # ==========================================================
    # PENTIFICATION
    # ==========================================================

    elif scenario == "Pentification":

        # Hypothèse simplifiée :
        # hausse de 50 pb sur la partie longue
        # et stabilité de la partie courte.

        variation_taux = 50 / 10000

        impact = (
            -sensibilite
            * variation_taux
            * 100
        )

        scenario_pb = "Scénario non parallèle"

        interpretation_scenario = (
            "Hausse de la partie longue de la courbe "
            "avec stabilité relative de la partie courte."
        )

    # ==========================================================
    # RÉSULTATS
    # ==========================================================

    st.divider()

    c1, c2, c3 = st.columns(3)

    c1.metric(
        "Sensibilité",
        f"{sensibilite:.2f}"
    )

    c2.metric(
        "Scénario",
        scenario_pb
    )

    c3.metric(
        "Impact estimé",
        f"{impact:.2f}%"
    )

    st.caption(
        interpretation_scenario
    )

    # ==========================================================
    # GRAPHIQUE
    # ==========================================================

    st.divider()

    graphique = pd.DataFrame({

        "Situation": [
            "Avant",
            "Après stress",
        ],

        "Impact (%)": [
            0,
            impact,
        ],

    })

    st.subheader(
        "📊 Impact estimé du scénario"
    )

    st.bar_chart(
        graphique.set_index(
            "Situation"
        )
    )

    # ==========================================================
    # INTERPRÉTATION
    # ==========================================================

    st.divider()

    st.subheader(
        "💬 Interprétation"
    )

    if impact > 0:

        st.success(
            f"""
            **Impact positif estimé de {impact:.2f}%**.

            Le scénario de baisse des taux est favorable à la valeur
            du portefeuille obligataire.
            """
        )

    elif impact > -1:

        st.success(
            f"""
            **Impact estimé : {impact:.2f}%**.

            Le fonds présente une sensibilité relativement faible
            au scénario sélectionné.
            """
        )

    elif impact > -3:

        st.warning(
            f"""
            **Impact estimé : {impact:.2f}%**.

            Le fonds présente une sensibilité modérée au scénario
            de taux sélectionné.
            """
        )

    else:

        st.error(
            f"""
            **Impact estimé : {impact:.2f}%**.

            Le fonds présente une sensibilité élevée au scénario
            de taux sélectionné.
            """
        )

    # ==========================================================
    # FONDS OMLT LES PLUS RÉSISTANTS
    # ==========================================================

    st.divider()

    st.subheader(
        "🛡️ Fonds OMLT les plus résistants à une hausse de +100 pb"
    )

    stress_omlt = fonds_obligataires[
        fonds_obligataires["Classification"] == "OMLT"
    ].copy()

    stress_omlt["Impact_100pb (%)"] = (
        -stress_omlt["Sensibilité"]
    )

    stress_omlt = stress_omlt.sort_values(
        "Impact_100pb (%)",
        ascending=False
    )

    st.dataframe(
        stress_omlt[
            [
                "OPCVM",
                "Classification",
                "Sensibilité",
                "Impact_100pb (%)",
            ]
        ].head(20),
        hide_index=True,
        use_container_width=True,
    )

    # ==========================================================
    # FONDS OCT LES PLUS RÉSISTANTS
    # ==========================================================

    st.subheader(
        "🛡️ Fonds OCT les plus résistants à une hausse de +100 pb"
    )

    stress_oct = fonds_obligataires[
        fonds_obligataires["Classification"] == "OCT"
    ].copy()

    stress_oct["Impact_100pb (%)"] = (
        -stress_oct["Sensibilité"]
    )

    stress_oct = stress_oct.sort_values(
        "Impact_100pb (%)",
        ascending=False
    )

    st.dataframe(
        stress_oct[
            [
                "OPCVM",
                "Classification",
                "Sensibilité",
                "Impact_100pb (%)",
            ]
        ].head(20),
        hide_index=True,
        use_container_width=True,
    )
# ==========================================================
# RAPPORT
# ==========================================================

# ==========================================================
# RAPPORT
# ==========================================================

elif page == "Rapport":

    import io
    import re

    from openpyxl import Workbook
    from openpyxl.styles import (
        Font,
        PatternFill,
        Alignment,
        Border,
        Side
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo


    # ==========================================================
    # OUTILS
    # ==========================================================

    def nettoyer_nom_feuille(nom):
        """
        Nettoie le nom d'une feuille Excel.
        Excel interdit certains caractères et limite à 31 caractères.
        """

        nom = str(nom)

        nom = re.sub(
            r'[\[\]\:\*\?\/\\]',
            '',
            nom
        )

        nom = nom.strip()

        if not nom:
            nom = "Feuille"

        return nom[:31]


    def nom_feuille_unique(wb, nom):
        """
        Garantit un nom de feuille unique.
        """

        nom = nettoyer_nom_feuille(nom)

        nom_initial = nom
        compteur = 1

        while nom in wb.sheetnames:

            suffixe = f"_{compteur}"

            nom = (
                nom_initial[:31 - len(suffixe)]
                + suffixe
            )

            compteur += 1

        return nom


    # ==========================================================
    # EXPORT CSV
    # ==========================================================

    def preparer_csv(
        titre,
        description,
        dataframe
    ):

        contenu = (
            f"{titre}\n"
            f"{description}\n\n"
        )

        contenu += dataframe.to_csv(
            index=False,
            encoding="utf-8-sig"
        )

        return contenu


    # ==========================================================
    # STYLES EXCEL
    # ==========================================================

    BLEU_FONCE = "1F4E78"
    BLEU = "5B9BD5"
    BLEU_CLAIR = "D9EAF7"
    GRIS = "F2F2F2"
    BLANC = "FFFFFF"
    VERT = "70AD47"
    ROUGE = "C00000"
    ORANGE = "ED7D31"

    BORDURE = Side(
        style="thin",
        color="D9E1F2"
    )


    # ==========================================================
    # AJUSTER LARGEUR
    # ==========================================================

    def ajuster_largeur(ws):

        for colonne in ws.columns:

            longueur_max = 0

            lettre = get_column_letter(
                colonne[0].column
            )

            for cellule in colonne:

                try:

                    longueur = len(
                        str(cellule.value)
                    )

                    longueur_max = max(
                        longueur_max,
                        longueur
                    )

                except Exception:

                    pass

            ws.column_dimensions[
                lettre
            ].width = min(
                max(
                    longueur_max + 3,
                    12
                ),
                35
            )


    # ==========================================================
    # ECRIRE UN TITRE DE SECTION
    # ==========================================================

    def ecrire_section(
        ws,
        ligne,
        titre,
        nombre_colonnes
    ):

        derniere_colonne = get_column_letter(
            max(nombre_colonnes, 1)
        )

        ws.merge_cells(
            f"A{ligne}:{derniere_colonne}{ligne}"
        )

        cellule = ws[f"A{ligne}"]

        cellule.value = titre

        cellule.font = Font(
            bold=True,
            size=12,
            color=BLANC
        )

        cellule.fill = PatternFill(
            "solid",
            fgColor=BLEU
        )

        cellule.alignment = Alignment(
            horizontal="left",
            vertical="center"
        )

        ws.row_dimensions[
            ligne
        ].height = 24


    # ==========================================================
    # ECRIRE UN DATAFRAME DANS EXCEL
    # ==========================================================

    def ecrire_dataframe(
        ws,
        dataframe,
        ligne_depart,
        nom_table=None
    ):

        dataframe = dataframe.copy()

        if dataframe.empty:

            ws.cell(
                row=ligne_depart,
                column=1
            ).value = "Aucune donnée disponible."

            return ligne_depart + 2


        # ------------------------------------------------------
        # EN-TÊTES
        # ------------------------------------------------------

        for colonne_index, colonne in enumerate(
            dataframe.columns,
            start=1
        ):

            cellule = ws.cell(
                row=ligne_depart,
                column=colonne_index
            )

            cellule.value = colonne

            cellule.font = Font(
                bold=True,
                color=BLANC
            )

            cellule.fill = PatternFill(
                "solid",
                fgColor=BLEU
            )

            cellule.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

            cellule.border = Border(
                bottom=BORDURE
            )


        # ------------------------------------------------------
        # DONNÉES
        # ------------------------------------------------------

        for ligne_index, ligne in enumerate(
            dataframe.itertuples(index=False),
            start=ligne_depart + 1
        ):

            for colonne_index, valeur in enumerate(
                ligne,
                start=1
            ):

                cellule = ws.cell(
                    row=ligne_index,
                    column=colonne_index
                )

                if pd.isna(valeur):

                    cellule.value = None

                else:

                    cellule.value = valeur

                cellule.border = Border(
                    bottom=BORDURE
                )

                cellule.alignment = Alignment(
                    vertical="center",
                    wrap_text=True
                )


        derniere_ligne = (
            ligne_depart
            + len(dataframe)
        )

        derniere_colonne = get_column_letter(
            len(dataframe.columns)
        )


        # ------------------------------------------------------
        # FORMAT DES NOMBRES
        # ------------------------------------------------------

        for colonne_index, colonne in enumerate(
            dataframe.columns,
            start=1
        ):

            nom_colonne = str(colonne)

            lettre = get_column_letter(
                colonne_index
            )

            # Pourcentages
            if (
                "%"
                in nom_colonne
                or "Performance"
                in nom_colonne
                or "Volatilité"
                in nom_colonne
                or "Régularité"
                in nom_colonne
            ):

                for ligne in range(
                    ligne_depart + 1,
                    derniere_ligne + 1
                ):

                    ws[
                        f"{lettre}{ligne}"
                    ].number_format = "0.00"


            # Scores / ratios
            elif (
                "Score"
                in nom_colonne
                or "Ratio"
                in nom_colonne
                or "Rang"
                in nom_colonne
            ):

                for ligne in range(
                    ligne_depart + 1,
                    derniere_ligne + 1
                ):

                    ws[
                        f"{lettre}{ligne}"
                    ].number_format = "0.00"


        # ------------------------------------------------------
        # TABLEAU EXCEL
        # ------------------------------------------------------

        if nom_table:

            nom_table = re.sub(
                r'[^A-Za-z0-9_]',
                '',
                str(nom_table)
            )

            if not nom_table:

                nom_table = "Tableau"

            reference = (
                f"A{ligne_depart}:"
                f"{derniere_colonne}{derniere_ligne}"
            )

            try:

                tableau = Table(
                    displayName=nom_table,
                    ref=reference
                )

                style = TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False
                )

                tableau.tableStyleInfo = style

                ws.add_table(
                    tableau
                )

            except Exception:

                pass


        return derniere_ligne + 2


    # ==========================================================
    # PREPARATION DES DONNEES D'UNE CATEGORIE
    # ==========================================================

    def preparer_donnees_categorie(
        categorie
    ):

        scoring_cat = table_scoring[
            table_scoring[
                "Classification"
            ] == categorie
        ].copy()


        risque_cat = table_risque[
            table_risque[
                "Classification"
            ] == categorie
        ].copy()


        # ------------------------------------------------------
        # 1. MEILLEURES PERFORMANCES
        # ------------------------------------------------------

        meilleures = (
            scoring_cat
            .sort_values(
                "Perf_YTD_calculee_%",
                ascending=False
            )
            .head(10)
            .copy()
        )

        meilleures_affichage = meilleures[
            [
                "OPCVM",
                "Perf_YTD_calculee_%",
                "Score_Global"
            ]
        ].copy()


        # ------------------------------------------------------
        # 2. PLUS FORTES BAISSES
        # ------------------------------------------------------

        baisses = (
            scoring_cat
            .sort_values(
                "Perf_YTD_calculee_%",
                ascending=True
            )
            .head(10)
            .copy()
        )

        baisses_affichage = baisses[
            [
                "OPCVM",
                "Perf_YTD_calculee_%",
                "Score_Global"
            ]
        ].copy()


        # ------------------------------------------------------
        # 3. PLUS REGULIERS
        # ------------------------------------------------------

        reguliers = (
            risque_cat
            .sort_values(
                "Pct_mois_positifs",
                ascending=False
            )
            .head(10)
            .copy()
        )

        reguliers_affichage = reguliers[
            [
                "OPCVM",
                "Pct_mois_positifs"
            ]
        ].copy()


        # ------------------------------------------------------
        # 4. PLUS VOLATILS
        # ------------------------------------------------------

        volatils = (
            risque_cat
            .sort_values(
                "Volatilite_annualisee_%",
                ascending=False
            )
            .head(10)
            .copy()
        )

        volatils_affichage = volatils[
            [
                "OPCVM",
                "Volatilite_annualisee_%"
            ]
        ].copy()


        # ------------------------------------------------------
        # 5. PERFORMANCE AJUSTEE DU RISQUE
        # ------------------------------------------------------

        risque_adjusted = risque_cat[
            [
                "OPCVM",
                "Perf_YTD_calculee_%",
                "Volatilite_annualisee_%"
            ]
        ].copy()


        risque_adjusted[
            "Perf_YTD_calculee_%"
        ] = pd.to_numeric(
            risque_adjusted[
                "Perf_YTD_calculee_%"
            ],
            errors="coerce"
        )


        risque_adjusted[
            "Volatilite_annualisee_%"
        ] = pd.to_numeric(
            risque_adjusted[
                "Volatilite_annualisee_%"
            ],
            errors="coerce"
        )


        risque_adjusted = (
            risque_adjusted
            .dropna(
                subset=[
                    "Perf_YTD_calculee_%",
                    "Volatilite_annualisee_%"
                ]
            )
        )


        risque_adjusted = risque_adjusted[
            risque_adjusted[
                "Volatilite_annualisee_%"
            ] > 0
        ]


        risque_adjusted[
            "Performance / Volatilité"
        ] = (
            risque_adjusted[
                "Perf_YTD_calculee_%"
            ]
            /
            risque_adjusted[
                "Volatilite_annualisee_%"
            ]
        )


        risque_adjusted = (
            risque_adjusted
            .sort_values(
                "Performance / Volatilité",
                ascending=False
            )
            .head(10)
        )


        risque_affichage = risque_adjusted[
            [
                "OPCVM",
                "Perf_YTD_calculee_%",
                "Volatilite_annualisee_%",
                "Performance / Volatilité"
            ]
        ].copy()


        # ------------------------------------------------------
        # 6. SYNTHESE
        # ------------------------------------------------------

        performance_moyenne = pd.to_numeric(
            scoring_cat[
                "Perf_YTD_calculee_%"
            ],
            errors="coerce"
        ).mean()


        volatilite_moyenne = pd.to_numeric(
            risque_cat[
                "Volatilite_annualisee_%"
            ],
            errors="coerce"
        ).mean()


        regularite_moyenne = pd.to_numeric(
            risque_cat[
                "Pct_mois_positifs"
            ],
            errors="coerce"
        ).mean()


        nombre_fonds = len(
            scoring_cat
        )


        synthese = pd.DataFrame({

            "Catégorie": [
                categorie
            ],

            "Nombre de fonds": [
                nombre_fonds
            ],

            "Performance moyenne YTD (%)": [
                performance_moyenne
            ],

            "Volatilité moyenne (%)": [
                volatilite_moyenne
            ],

            "Régularité moyenne (%)": [
                regularite_moyenne
            ]

        })


        return {

            "meilleures": meilleures_affichage,

            "baisses": baisses_affichage,

            "reguliers": reguliers_affichage,

            "volatils": volatils_affichage,

            "risque": risque_affichage,

            "synthese": synthese

        }


    # ==========================================================
    # GENERATION DU LIVRABLE 2
    # ==========================================================

    def generer_livrable2():

        output = io.BytesIO()

        wb = Workbook()

        # Supprimer la feuille par défaut
        ws = wb.active

        ws.title = "Synthèse"


        # ======================================================
        # FEUILLE SYNTHESE
        # ======================================================

        ws = wb["Synthèse"]

        # ------------------------------------------------------
        # TITRE
        # ------------------------------------------------------

        ws.merge_cells(
            "A1:E1"
        )

        ws["A1"] = (
            "TABLEAU DE BORD OPCVM MAROCAINS"
        )

        ws["A1"].font = Font(
            bold=True,
            size=18,
            color=BLANC
        )

        ws["A1"].fill = PatternFill(
            "solid",
            fgColor=BLEU_FONCE
        )

        ws["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        ws.row_dimensions[1].height = 32


        # ------------------------------------------------------
        # DESCRIPTION
        # ------------------------------------------------------

        ws.merge_cells(
            "A2:E2"
        )

        ws["A2"] = (
            "Synthèse quantitative des OPCVM marocains "
            "par catégorie : performance, risque, régularité "
            "et performance ajustée du risque."
        )

        ws["A2"].font = Font(
            italic=True,
            color="404040"
        )

        ws["A2"].fill = PatternFill(
            "solid",
            fgColor=BLEU_CLAIR
        )

        ws["A2"].alignment = Alignment(
            wrap_text=True,
            vertical="center"
        )

        ws.row_dimensions[2].height = 40


        # ------------------------------------------------------
        # INDICATEURS GENERAUX
        # ------------------------------------------------------

        ws["A4"] = "INDICATEURS GÉNÉRAUX"

        ws["A4"].font = Font(
            bold=True,
            size=13,
            color=BLANC
        )

        ws["A4"].fill = PatternFill(
            "solid",
            fgColor=BLEU
        )


        indicateurs = pd.DataFrame({

            "Indicateur": [
                "Nombre de fonds",
                "Nombre de catégories",
                "Nombre de sociétés de gestion",
                "Nombre de semaines analysées"
            ],

            "Valeur": [
                table_reference[
                    "CODE ISIN"
                ].nunique(),

                table_reference[
                    "Classification"
                ].nunique(),

                table_reference[
                    "Société de Gestion"
                ].nunique(),

                len(vl)
            ]

        })


        ecrire_dataframe(
            ws,
            indicateurs,
            5,
            "TableauIndicateurs"
        )


        # ------------------------------------------------------
        # REPARTITION DES CATEGORIES
        # ------------------------------------------------------

        ligne = 12

        ecrire_section(
            ws,
            ligne,
            "RÉPARTITION DES OPCVM PAR CATÉGORIE",
            2
        )

        repartition = (
            table_reference[
                "Classification"
            ]
            .value_counts()
            .rename_axis(
                "Classification"
            )
            .reset_index(
                name="Nombre de fonds"
            )
        )

        ecrire_dataframe(
            ws,
            repartition,
            ligne + 1,
            "TableauRepartition"
        )


        # ------------------------------------------------------
        # SYNTHESE DE TOUTES LES CATEGORIES
        # ------------------------------------------------------

        syntheses = []

        categories_export = sorted(
            table_scoring[
                "Classification"
            ]
            .dropna()
            .unique()
        )


        for categorie in categories_export:

            donnees = preparer_donnees_categorie(
                categorie
            )

            syntheses.append(
                donnees["synthese"]
            )


        if syntheses:

            synthese_globale = pd.concat(
                syntheses,
                ignore_index=True
            )

        else:

            synthese_globale = pd.DataFrame()


        ligne_synthese = 25

        ecrire_section(
            ws,
            ligne_synthese,
            "SYNTHÈSE PAR CATÉGORIE",
            5
        )

        ecrire_dataframe(
            ws,
            synthese_globale,
            ligne_synthese + 1,
            "TableauSynthese"
        )


        ws.freeze_panes = "A5"

        ajuster_largeur(ws)


        # ======================================================
        # FEUILLES PAR CATEGORIE
        # ======================================================

        for categorie in categories_export:

            donnees = preparer_donnees_categorie(
                categorie
            )


            nom_feuille = nom_feuille_unique(
                wb,
                categorie
            )

            ws = wb.create_sheet(
                title=nom_feuille
            )


            # --------------------------------------------------
            # TITRE
            # --------------------------------------------------

            ws.merge_cells(
                "A1:D1"
            )

            ws["A1"] = (
                f"ANALYSE OPCVM – {categorie}"
            )

            ws["A1"].font = Font(
                bold=True,
                size=17,
                color=BLANC
            )

            ws["A1"].fill = PatternFill(
                "solid",
                fgColor=BLEU_FONCE
            )

            ws["A1"].alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            ws.row_dimensions[1].height = 30


            # --------------------------------------------------
            # DESCRIPTION
            # --------------------------------------------------

            ws.merge_cells(
                "A2:D2"
            )

            ws["A2"] = (
                "Classements calculés séparément pour cette "
                "catégorie afin de comparer des OPCVM de "
                "profil homogène."
            )

            ws["A2"].font = Font(
                italic=True,
                color="404040"
            )

            ws["A2"].fill = PatternFill(
                "solid",
                fgColor=BLEU_CLAIR
            )

            ws["A2"].alignment = Alignment(
                wrap_text=True,
                vertical="center"
            )

            ws.row_dimensions[2].height = 35


            # --------------------------------------------------
            # SYNTHESE
            # --------------------------------------------------

            ligne = 4

            ecrire_section(
                ws,
                ligne,
                "SYNTHÈSE DE LA CATÉGORIE",
                5
            )

            ligne = ecrire_dataframe(
                ws,
                donnees["synthese"],
                ligne + 1,
                "TableauSyntheseCat"
            )


            # --------------------------------------------------
            # MEILLEURES PERFORMANCES
            # --------------------------------------------------

            ecrire_section(
                ws,
                ligne,
                "🏆 MEILLEURES PERFORMANCES",
                3
            )

            ligne = ecrire_dataframe(
                ws,
                donnees["meilleures"],
                ligne + 1,
                "TableauPerformances"
            )


            # --------------------------------------------------
            # PLUS FORTES BAISSES
            # --------------------------------------------------

            ecrire_section(
                ws,
                ligne,
                "📉 PLUS FORTES BAISSES",
                3
            )

            ligne = ecrire_dataframe(
                ws,
                donnees["baisses"],
                ligne + 1,
                "TableauBaisses"
            )


            # --------------------------------------------------
            # PLUS REGULIERS
            # --------------------------------------------------

            ecrire_section(
                ws,
                ligne,
                "🎯 FONDS LES PLUS RÉGULIERS",
                2
            )

            ligne = ecrire_dataframe(
                ws,
                donnees["reguliers"],
                ligne + 1,
                "TableauReguliers"
            )


            # --------------------------------------------------
            # PLUS VOLATILS
            # --------------------------------------------------

            ecrire_section(
                ws,
                ligne,
                "📈 FONDS LES PLUS VOLATILS",
                2
            )

            ligne = ecrire_dataframe(
                ws,
                donnees["volatils"],
                ligne + 1,
                "TableauVolatils"
            )


            # --------------------------------------------------
            # PERFORMANCE AJUSTEE DU RISQUE
            # --------------------------------------------------

            ecrire_section(
                ws,
                ligne,
                "⚖️ MEILLEURES PERFORMANCES AJUSTÉES DU RISQUE",
                4
            )

            ligne = ecrire_dataframe(
                ws,
                donnees["risque"],
                ligne + 1,
                "TableauRisque"
            )


            # --------------------------------------------------
            # MISE EN PAGE
            # --------------------------------------------------

            ws.freeze_panes = "A5"

            ajuster_largeur(ws)


        # ======================================================
        # EXPORT
        # ======================================================

        wb.save(
            output
        )

        output.seek(0)

        return output.getvalue()


    # ==========================================================
    # TITRE DE LA PAGE
    # ==========================================================

    st.title(
        "📄 Rapport de synthèse"
    )

    st.caption(
        "Synthèse quantitative des OPCVM marocains"
    )

    st.divider()


    # ==========================================================
    # INDICATEURS GENERAUX
    # ==========================================================

    c1, c2, c3, c4 = st.columns(4)


    c1.metric(
        "Nombre de fonds",
        table_reference[
            "CODE ISIN"
        ].nunique()
    )


    c2.metric(
        "Catégories",
        table_reference[
            "Classification"
        ].nunique()
    )


    c3.metric(
        "Sociétés de gestion",
        table_reference[
            "Société de Gestion"
        ].nunique()
    )


    c4.metric(
        "Semaines analysées",
        len(vl)
    )


    # ==========================================================
    # REPARTITION DES CATEGORIES
    # ==========================================================

    st.subheader(
        "📊 Répartition des catégories"
    )


    repartition = (
        table_reference[
            "Classification"
        ]
        .value_counts()
        .rename_axis(
            "Classification"
        )
        .reset_index(
            name="Nombre de fonds"
        )
    )


    st.dataframe(
        repartition.style
        .background_gradient(
            subset=[
                "Nombre de fonds"
            ],
            cmap="Blues"
        )
        .format({
            "Nombre de fonds": "{:,.0f}"
        }),
        hide_index=True,
        use_container_width=True
    )


    st.bar_chart(
        repartition.set_index(
            "Classification"
        )
    )


    # ==========================================================
    # CATEGORIES
    # ==========================================================

    categories = sorted(
        table_scoring[
            "Classification"
        ]
        .dropna()
        .unique()
    )


    st.divider()

    st.header(
        "📚 Analyse détaillée par catégorie"
    )

    st.caption(
        "Les classements sont calculés séparément pour chaque "
        "catégorie afin de comparer les OPCVM à des fonds "
        "de profil homogène."
    )


    # ==========================================================
    # AFFICHAGE STREAMLIT
    # ==========================================================

    for categorie in categories:

        st.divider()

        st.header(
            f"📌 {categorie}"
        )


        donnees = preparer_donnees_categorie(
            categorie
        )


        # ======================================================
        # MEILLEURES PERFORMANCES
        # ======================================================

        st.subheader(
            "🏆 Meilleures performances"
        )

        st.dataframe(
            donnees[
                "meilleures"
            ].style
            .background_gradient(
                subset=[
                    "Perf_YTD_calculee_%"
                ],
                cmap="RdYlGn"
            )
            .format({
                "Perf_YTD_calculee_%": "{:.2f}%",
                "Score_Global": "{:.1f}"
            }),
            hide_index=True,
            use_container_width=True
        )


        # ======================================================
        # PLUS FORTES BAISSES
        # ======================================================

        st.subheader(
            "📉 Plus fortes baisses"
        )

        st.dataframe(
            donnees[
                "baisses"
            ].style
            .background_gradient(
                subset=[
                    "Perf_YTD_calculee_%"
                ],
                cmap="RdYlGn_r"
            )
            .format({
                "Perf_YTD_calculee_%": "{:.2f}%",
                "Score_Global": "{:.1f}"
            }),
            hide_index=True,
            use_container_width=True
        )


        # ======================================================
        # PLUS REGULIERS
        # ======================================================

        st.subheader(
            "🎯 Fonds les plus réguliers"
        )

        st.dataframe(
            donnees[
                "reguliers"
            ].style
            .background_gradient(
                subset=[
                    "Pct_mois_positifs"
                ],
                cmap="Greens"
            )
            .format({
                "Pct_mois_positifs": "{:.1f}%"
            }),
            hide_index=True,
            use_container_width=True
        )


        # ======================================================
        # PLUS VOLATILS
        # ======================================================

        st.subheader(
            "📈 Fonds les plus volatils"
        )

        st.dataframe(
            donnees[
                "volatils"
            ].style
            .background_gradient(
                subset=[
                    "Volatilite_annualisee_%"
                ],
                cmap="OrRd"
            )
            .format({
                "Volatilite_annualisee_%": "{:.2f}%"
            }),
            hide_index=True,
            use_container_width=True
        )


        # ======================================================
        # PERFORMANCE AJUSTEE DU RISQUE
        # ======================================================

        st.subheader(
            "⚖️ Meilleures performances ajustées du risque"
        )

        st.dataframe(
            donnees[
                "risque"
            ].style
            .background_gradient(
                subset=[
                    "Performance / Volatilité"
                ],
                cmap="RdYlGn"
            )
            .format({
                "Perf_YTD_calculee_%": "{:.2f}%",
                "Volatilite_annualisee_%": "{:.2f}%",
                "Performance / Volatilité": "{:.2f}"
            }),
            hide_index=True,
            use_container_width=True
        )

        st.caption(
            "Performance ajustée du risque = "
            "Performance YTD / Volatilité annualisée."
        )


        # ======================================================
        # SYNTHESE
        # ======================================================

        st.subheader(
            "📊 Synthèse de la catégorie"
        )

        synthese = donnees[
            "synthese"
        ]

        st.dataframe(
            synthese.style
            .background_gradient(
                subset=[
                    "Performance moyenne YTD (%)"
                ],
                cmap="RdYlGn"
            )
            .format({
                "Performance moyenne YTD (%)": "{:.2f}",
                "Volatilité moyenne (%)": "{:.2f}",
                "Régularité moyenne (%)": "{:.1f}"
            }),
            hide_index=True,
            use_container_width=True
        )


    # ==========================================================
    # RAPPORTS EXCEL
    # ==========================================================

    st.divider()

    st.subheader("📊 Rapports Excel")

    st.caption(
        "Téléchargez les deux livrables principaux de la plateforme."
    )


    # ==========================================================
    # FONCTION : GÉNÉRER LE LIVRABLE 1
    # ==========================================================

    def generer_livrable1():

        output = io.BytesIO()

        wb = Workbook()

        # Supprimer la feuille créée automatiquement
        ws = wb.active
        ws.title = "Base OPCVM"


        # ======================================================
        # COULEURS
        # ======================================================

        bleu_fonce = "1F4E78"
        bleu = "5B9BD5"
        bleu_clair = "D9EAF7"
        blanc = "FFFFFF"

        bordure = Side(
            style="thin",
            color="D9E1F2"
        )


        # ======================================================
        # FONCTION INTERNE : AJOUTER UN DATAFRAME
        # ======================================================

        def ajouter_dataframe(
            ws,
            dataframe,
            ligne_depart=1
        ):

            dataframe = dataframe.copy()

            # ----------------------------------------------
            # En-têtes
            # ----------------------------------------------

            for colonne_index, colonne in enumerate(
                dataframe.columns,
                start=1
            ):

                cellule = ws.cell(
                    row=ligne_depart,
                    column=colonne_index
                )

                cellule.value = colonne

                cellule.font = Font(
                    bold=True,
                    color=blanc
                )

                cellule.fill = PatternFill(
                    "solid",
                    fgColor=bleu
                )

                cellule.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

                cellule.border = Border(
                    bottom=bordure
                )


            # ----------------------------------------------
            # Données
            # ----------------------------------------------

            for ligne_index, ligne in enumerate(
                dataframe.itertuples(index=False),
                start=ligne_depart + 1
            ):

                for colonne_index, valeur in enumerate(
                    ligne,
                    start=1
                ):

                    cellule = ws.cell(
                        row=ligne_index,
                        column=colonne_index
                    )

                    if pd.isna(valeur):
                        cellule.value = None
                    else:
                        cellule.value = valeur

                    cellule.alignment = Alignment(
                        vertical="center",
                        wrap_text=True
                    )

                    cellule.border = Border(
                        bottom=bordure
                    )


            # ----------------------------------------------
            # Largeur des colonnes
            # ----------------------------------------------

            for colonne_index, colonne in enumerate(
                dataframe.columns,
                start=1
            ):

                lettre = get_column_letter(
                    colonne_index
                )

                longueur_max = len(
                    str(colonne)
                )

                for valeur in dataframe.iloc[
                    :, colonne_index - 1
                ]:

                    if pd.isna(valeur):
                        longueur = 0
                    else:
                        longueur = len(
                            str(valeur)
                        )

                    longueur_max = max(
                        longueur_max,
                        longueur
                    )

                ws.column_dimensions[
                    lettre
                ].width = min(
                    max(
                        longueur_max + 3,
                        12
                    ),
                    35
                )


            # ----------------------------------------------
            # Volet figé
            # ----------------------------------------------

            ws.freeze_panes = (
                f"A{ligne_depart + 1}"
            )


        # ======================================================
        # 1. BASE OPCVM
        # ======================================================

        base_opcvm = table_reference.copy()

        ws = wb["Base OPCVM"]

        ws["A1"] = (
            "LIVRABLE 1 – BASE DE DONNÉES OPCVM"
        )

        ws["A1"].font = Font(
            bold=True,
            size=15,
            color=blanc
        )

        ws["A1"].fill = PatternFill(
            "solid",
            fgColor=bleu_fonce
        )

        ws["A1"].alignment = Alignment(
            horizontal="center"
        )

        # Fusion du titre
        nombre_colonnes = max(
            len(base_opcvm.columns),
            1
        )

        derniere_colonne = get_column_letter(
            nombre_colonnes
        )

        ws.merge_cells(
            f"A1:{derniere_colonne}1"
        )

        ws.row_dimensions[1].height = 30


        ws["A2"] = (
            "Base structurée des OPCVM marocains : "
            "identification, classification et informations "
            "de référence des fonds."
        )

        ws["A2"].fill = PatternFill(
            "solid",
            fgColor=bleu_clair
        )

        ws["A2"].alignment = Alignment(
            wrap_text=True
        )

        ws.merge_cells(
            f"A2:{derniere_colonne}2"
        )

        ws.row_dimensions[2].height = 35


        ajouter_dataframe(
            ws,
            base_opcvm,
            ligne_depart=4
        )


        # ======================================================
        # 2. PERFORMANCES OPCVM
        # ======================================================

        if "table_scoring" in globals():

            ws_perf = wb.create_sheet(
                "Performances"
            )

            performances = table_scoring.copy()

            ajouter_dataframe(
                ws_perf,
                performances,
                ligne_depart=1
            )


        # ======================================================
        # 3. RISQUES OPCVM
        # ======================================================

        if "table_risque" in globals():

            ws_risque = wb.create_sheet(
                "Risques"
            )

            risques = table_risque.copy()

            ajouter_dataframe(
                ws_risque,
                risques,
                ligne_depart=1
            )


        # ======================================================
        # 4. BENCHMARKS
        # ======================================================

        # On récupère le tableau des benchmarks s'il existe
        benchmarks = None

        if "base_benchmarks" in globals():

            benchmarks = base_benchmarks.copy()

        elif "table_benchmark" in globals():

            benchmarks = table_benchmark.copy()

        elif "benchmarks" in globals():

            if isinstance(
                globals()["benchmarks"],
                pd.DataFrame
            ):

                benchmarks = globals()[
                    "benchmarks"
                ].copy()


        if benchmarks is not None:

            ws_bench = wb.create_sheet(
                "Benchmarks"
            )

            ws_bench["A1"] = (
                "BENCHMARKS DE RÉFÉRENCE"
            )

            ws_bench["A1"].font = Font(
                bold=True,
                size=15,
                color=blanc
            )

            ws_bench["A1"].fill = PatternFill(
                "solid",
                fgColor=bleu_fonce
            )

            ws_bench["A1"].alignment = Alignment(
                horizontal="center"
            )

            nb_colonnes = max(
                len(benchmarks.columns),
                1
            )

            derniere_colonne = get_column_letter(
                nb_colonnes
            )

            ws_bench.merge_cells(
                f"A1:{derniere_colonne}1"
            )

            ws_bench["A2"] = (
                "Données des indices et références "
                "utilisés pour l'analyse comparative "
                "des OPCVM."
            )

            ws_bench["A2"].fill = PatternFill(
                "solid",
                fgColor=bleu_clair
            )

            ws_bench["A2"].alignment = Alignment(
                wrap_text=True
            )

            ws_bench.merge_cells(
                f"A2:{derniere_colonne}2"
            )

            ajouter_dataframe(
                ws_bench,
                benchmarks,
                ligne_depart=4
            )


        # ======================================================
        # 5. COMPARAISON OPCVM / BENCHMARK
        # ======================================================

        if benchmarks is not None:

            ws_comp = wb.create_sheet(
                "Comparaison Benchmark"
            )

            # ----------------------------------------------
            # Préparation
            # ----------------------------------------------

            comparaison = None

            # Chercher les colonnes de performance
            colonne_opcvm = None
            colonne_benchmark = None

            for col in benchmarks.columns:

                nom = str(col).lower()

                if (
                    "performance" in nom
                    or "perf" in nom
                    or "rendement" in nom
                ):

                    colonne_benchmark = col
                    break


            if (
                "OPCVM" in table_scoring.columns
                and "Perf_YTD_calculee_%" in table_scoring.columns
            ):

                comparaison = table_scoring[
                    [
                        "OPCVM",
                        "Classification",
                        "Perf_YTD_calculee_%"
                    ]
                ].copy()

                comparaison.rename(
                    columns={
                        "Perf_YTD_calculee_%":
                            "Performance OPCVM YTD (%)"
                    },
                    inplace=True
                )


            # ----------------------------------------------
            # Si aucune structure compatible
            # ----------------------------------------------

            if comparaison is None:

                comparaison = pd.DataFrame({
                    "Information": [
                        "Aucune comparaison benchmark disponible"
                    ]
                })


            # ----------------------------------------------
            # Ajouter colonne benchmark si possible
            # ----------------------------------------------

            if (
                comparaison is not None
                and colonne_benchmark is not None
            ):

                # On ne force pas une jointure incorrecte.
                # Le tableau de benchmark reste disponible
                # dans sa feuille dédiée.

                comparaison[
                    "Benchmark de référence"
                ] = "À associer selon la catégorie"


            # ----------------------------------------------
            # Titre
            # ----------------------------------------------

            ws_comp["A1"] = (
                "COMPARAISON DES OPCVM AVEC LES BENCHMARKS"
            )

            ws_comp["A1"].font = Font(
                bold=True,
                size=15,
                color=blanc
            )

            ws_comp["A1"].fill = PatternFill(
                "solid",
                fgColor=bleu_fonce
            )

            ws_comp["A1"].alignment = Alignment(
                horizontal="center"
            )

            nb_colonnes = max(
                len(comparaison.columns),
                1
            )

            derniere_colonne = get_column_letter(
                nb_colonnes
            )

            ws_comp.merge_cells(
                f"A1:{derniere_colonne}1"
            )

            ws_comp["A2"] = (
                "Tableau destiné à comparer la performance "
                "des OPCVM avec leur indice ou référence "
                "de marché approprié."
            )

            ws_comp["A2"].fill = PatternFill(
                "solid",
                fgColor=bleu_clair
            )

            ws_comp["A2"].alignment = Alignment(
                wrap_text=True
            )

            ws_comp.merge_cells(
                f"A2:{derniere_colonne}2"
            )


            ajouter_dataframe(
                ws_comp,
                comparaison,
                ligne_depart=4
            )


        # ======================================================
        # 6. CLASSIFICATION
        # ======================================================

        ws_class = wb.create_sheet(
            "Classification"
        )

        classification = (
            table_reference[
                [
                    "CODE ISIN",
                    "OPCVM",
                    "Classification",
                    "Société de Gestion"
                ]
            ]
            .drop_duplicates()
            .copy()
        )

        ajouter_dataframe(
            ws_class,
            classification,
            ligne_depart=1
        )


        # ======================================================
        # 7. STYLE DES FEUILLES
        # ======================================================

        for feuille in wb.worksheets:

            feuille.sheet_view.showGridLines = False

            feuille.auto_filter.ref = (
                feuille.dimensions
            )


        # ======================================================
        # EXPORT
        # ======================================================

        wb.save(output)

        output.seek(0)

        return output.getvalue()


    # ==========================================================
    # LIVRABLE 1 – BASE DE DONNÉES OPCVM
    # ==========================================================

    st.markdown(
        "### 📗 Livrable 1 – Base de données OPCVM"
    )

    livrable1 = generer_livrable1()

    st.download_button(
        label="📥 Télécharger le Livrable 1",
        data=livrable1,
        file_name="livrable 1.xlsx",
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        key="livrable1_excel"
    )

    st.caption(
        "Base de données structurée des OPCVM comprenant "
        "les informations de référence, les performances, "
        "les risques, la classification et les benchmarks "
        "de référence."
    )


    # ==========================================================
    # LIVRABLE 2 – TABLEAU DE BORD OPCVM
    # ==========================================================

    st.markdown(
        "### 📊 Livrable 2 – Tableau de bord OPCVM"
    )

    st.download_button(
        label="📥 Télécharger le Livrable 2",
        data=generer_livrable2(),
        file_name="livrable 2.xlsx",
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        key="livrable2_excel"
    )

    st.caption(
        "Tableau de bord permettant de suivre les performances, "
        "les plus fortes baisses, la régularité, la volatilité, "
        "la performance ajustée du risque et les principaux "
        "indicateurs des OPCVM."
    )   
elif page == "Méthodologie":

    st.title("⚙️ Méthodologie")

    st.header("1. Sources des données")

    st.markdown(
        """
        Les données utilisées proviennent principalement de trois sources :

        - **ASFIM** : valeurs liquidatives, performances et caractéristiques des OPCVM
          (ISIN, société de gestion, classification, actif net, sensibilité, benchmark).
        - **Bank Al-Maghrib (BAM)** : MONIA, TMP et courbe des taux.
        - **BMCE Capital Bourse** : indices **MASI** et **MASI RB**.

        Les données ASFIM sont actualisées de manière hebdomadaire.
        """
    )

    st.header("2. Préparation des données")

    st.markdown(
        """
        Les données font l'objet d'un traitement préalable comprenant :

        - nettoyage et contrôle des valeurs liquidatives ;
        - gestion des valeurs manquantes et des anomalies ;
        - harmonisation des dates et des formats ;
        - normalisation des benchmarks ;
        - exclusion des fonds disposant de moins de **5 VL valides**.
        """
    )

    st.header("3. Indicateurs calculés")

    st.markdown(
        """
        L'analyse repose notamment sur :

        - **Performances** : 1 mois, 3 mois, 6 mois, 1 an, 3 ans, YTD et annualisée ;
        - **Risque** : volatilité annualisée, Maximum Drawdown et ratio rendement/volatilité ;
        - **Régularité** : pourcentage de mois positifs et écart-type des rendements mensuels ;
        - **Taille** : actif net du fonds.
        """
    )

    st.header("4. Benchmarks")

    st.markdown(
        """
        Les performances des OPCVM sont comparées, lorsque les données sont disponibles,
        à des références adaptées à leur catégorie :

        **MASI, MASI RB, MBI CT, MBI MT, MBI MLT, MBI GLOBAL, MONIA et TMP.**

        L'analyse permet notamment de mesurer l'écart de performance entre le fonds et son benchmark.
        """
    )

    st.header("5. Analyse de la courbe des taux")

    st.markdown(
        """
        La courbe des taux BAM est analysée selon trois segments :

        - **Court terme** : ≤ 2 ans
        - **Moyen terme** : 2 à 7 ans
        - **Long terme** : > 7 ans

        Les principaux indicateurs sont la moyenne des taux par segment, la pente
        **LT − CT**, les spreads et l'évolution de la courbe dans le temps.
        """
    )

    st.header("6. Stress tests")

    st.markdown(
        """
        Les OPCVM obligataires peuvent être soumis à différents scénarios de taux :

        - hausse ou baisse de **25, 50 et 100 points de base** ;
        - scénarios de **pentification** et d'**aplatissement**.

        L'impact est estimé à partir de la sensibilité du fonds selon :

        **ΔVL ≈ − Sensibilité × ΔTaux**
        """
    )

    st.header("7. Méthodologie du scoring")

    st.markdown(
        """
        Chaque fonds reçoit un **Score Global sur 100**.

        Les cinq composantes sont normalisées **par catégorie de fonds**, à l'aide de
        percentiles, puis agrégées selon les pondérations suivantes :

        - **Performance : 30 %**
        - **Volatilité : 25 %**
        - **Régularité : 20 %**
        - **Drawdown : 15 %**
        - **Taille : 10 %**

        Dans la version actuelle, le score de performance repose principalement sur la
        **performance YTD**.
        """
    )

    st.header("8. Hypothèses et limites")
    
    st.markdown(
    """
        Les résultats dépendent de la qualité et de la profondeur historique des données disponibles.
        
        Certaines limites doivent être prises en compte :

        - historique variable selon les fonds ;
        - benchmarks parfois indisponibles ou imparfaitement renseignés ;
        - sensibilité fournie par la société de gestion et pouvant rester approximative ;
        - stress tests fondés sur une approximation simplifiée ;
        - données hebdomadaires ne capturant pas les variations intra-semaine ;
        - frais d'entrée et de sortie non intégrés.

        Le scoring constitue un **outil d'aide à l'analyse et à la comparaison des OPCVM**.
        """
    )
    st.header("9. Outils utilisés")

    st.markdown(
        """
        Le projet a été développé avec **Python**, notamment :

        `Pandas` · `NumPy` · `Plotly` · `Streamlit` · `OpenPyXL` · `Requests`
        """
    )